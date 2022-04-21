#-------------------------------------------------------------------------------
# Name:        pull_input_data.py
# Purpose:     Functions to create TODO
# Author:      Dave Mcbey
# Created:     22/03/2020
# Description:
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'pull_input_data.py'
__version__ = '0.0.0'
__author__ = 's02dm4'

from pandas import read_excel
from openpyxl import load_workbook

REQD_SHEETS = {'Inputs4': 'Inputs4- Livestock', 'A1a': 'A1a. Soils and land use data',
                         'C1a':'C1a. Typical animal production', 'C1': 'C1. Change in animal production'}
                        #    'G9':'G9.Harvest and sowing months' }

def _read_sheet_vars(wb_obj, sheet_name, sheet_layout):
    '''
    read values from column B - parameter names are in column A
    '''
    if sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[sheet_name]
        sheet_vars = {}
        for parm in sheet_layout:
            cell_ref = sheet_layout[parm]
            cell = sheet_obj[cell_ref]
            sheet_vars[parm] = cell.value
    else:
        print('Sheet ' + sheet_name + ' not found in ')
        return None

    return sheet_vars

class ReadInputExcel(object, ):

    def __init__(self, form, xls_fname):
        '''
        read values and parameters from ORATOR inputs Excel file
        sheets read:    Inputs4- Livestock
                        A1a. Soils and land use data
                        C1. Change in animal production
                        C1a. Typical animal production
        '''
        func_name =  __prog__ +  ' oratorExcelDetail __init__'

        print('Loading: ' + xls_fname)
        wb_obj = load_workbook(xls_fname, data_only = True)    # gets values rather than formula
        all_sheets_flag = True
        for sheet_name in REQD_SHEETS.values():
            if sheet_name not in wb_obj.sheetnames:
                all_sheets_flag = False
                print('Missing sheet: ' + sheet_name)
                self.retcode = None

        if all_sheets_flag:
            # Farm information
            # ================
            sheet = REQD_SHEETS['C1']
            locat_layout = {'Region':'C18', 'Production':'F18', 'Climate':'H18', 'System':'K18'}
            self.farm_info = _read_sheet_vars(wb_obj, sheet, locat_layout)
            wb_obj.close()

            # livestock data for farm e.g. number of cattle, goats; feed type etc
            # ===================================================================
            sheet = REQD_SHEETS['Inputs4']
            print('Reading livestock from sheet: ' + sheet)
            data = read_excel(xls_fname, sheet_name=sheet, index_col=0, header=0, usecols=range(3,10),
                                                                                                skiprows=range(0,16))
            self.livestock_input = data.dropna(how='all')

            # Percent produced last harvest, typically 0.6 to 2.7            #
            # ===================================================
            sheet = REQD_SHEETS['A1a']
            print('Reading harvest data, percent produced, from sheet: ' + sheet)
            data = read_excel(xls_fname, sheet_name=sheet, usecols=range(40,45), skiprows=range(0,46))
            self.harvest_data = data.dropna(how='all')

            # land use data for 5 farms - typically 10 years of monthly data, integer values e.g. 4 is Maize
            # ==============================================================================================
            sheet = REQD_SHEETS['A1a']
            print('Reading last land use data from sheet: ' + sheet)
            data = read_excel(xls_fname, sheet_name=sheet, usecols=range(45,50), skiprows=range(0,46))
            self.last_land_use = data.dropna(how='all')

            # African animal production for different regions from Herrero (2016)
            # ===================================================================
            sheet = REQD_SHEETS['C1a']
            print('Reading Africa animal production data from sheet: ' + sheet)
            data = read_excel(xls_fname, header=0, sheet_name=sheet, usecols=range(1,15), skiprows=range(0,12))
            self.animal_prod = data.dropna(how='all')

            self.retcode = len(REQD_SHEETS)
