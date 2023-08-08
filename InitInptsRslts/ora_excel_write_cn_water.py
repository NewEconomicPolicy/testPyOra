"""
# -------------------------------------------------------------------------------
# Name:        ora_excel_write_cn_water.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Definitions:
#   spin_up
#
# Description:
#
#
# -------------------------------------------------------------------------------
# !/usr/bin/env python
"""
__prog__ = 'ora_excel_write_cn_water.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
from os.path import isfile, join, exists
from os import remove

from string import ascii_uppercase
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment
from pandas import DataFrame, ExcelWriter

from ora_lookup_df_fns import fetch_detail_from_varname

ALPHABET = list(ascii_uppercase)
WARN_STR = '*** Warning *** '
'''
26 colors for the 2010 Colour Alphabet Project which proposed 26 "letters" coded by colors 
stated to be near the maximum number reliably distinguishable. 
'''
COLORS_DISTING = {'Green': '2BCE48', 'Damson': '4C005C', 'Forest': '005C31', 'Sky': '5EF1F2', 'Khaki': '8F7C00',
                  'Lime': '9DCC00', 'Blue': '0075DC', 'Jade': '94FFB5', 'Violet': '740AFF', 'Caramel': '993F00',
                  'Turquoise': '00998F', 'Navy': '003380', 'Ebony': '191919', 'Quagmire': '426600', 'Iron': '808080',
                  'Wine': '990000', 'Mallow': 'C20088', 'Uranium': 'E0FF66', 'Amethyst': 'F0A3FF', 'Red': 'FF0010',
                  'Zinnia': 'FF5005', 'Pink': 'FFA8BB', 'Orpiment': 'FFA405', 'Honeydew': 'FFCC99', 'Yellow': 'FFFF00',
                  'Xanthin': 'FFFF80', 'White': 'FFFFFF'}

POOL_GROUPS = {'Active Pools': list(['pool_c_dpm', 'pool_c_rpm', 'pool_c_bio']),
               'Resistant Pools': list(['pool_c_hum', 'pool_c_iom', 'tot_soc_simul'])}
LINE_COLORS_POOL_GROUPS = list(['FF0000', '00FF00', '0000FF'])

LINE_COLORS_SUBAREAS = list(['Blue', 'Jade', 'Red', 'Honeydew', 'Yellow'])

CHANGE_VARS = {'carbon': list(['rate_mod', 'pool_c_dpm', 'pool_c_rpm', 'pool_c_bio', 'pool_c_hum', 'pool_c_iom',
                               'tot_soc_simul', 'co2_emiss']),
               'nitrogen': list(['soil_n_sply', 'no3_crop_dem', 'no3_nitrif', 'no3_leach', 'no3_denit',
                                 'nh4_crop_dem', 'nh4_volat']),
               'water': list(['wc_pwp', 'wat_soil', 'wc_fld_cap', 'wat_strss_indx', 'aet', 'irrig',
                              'wc_soil_irri_root_zone', 'aet_irri', 'wc_soil_irri', 'wat_drain', 'pcnt_c'])}

PREFERRED_LINE_WIDTH = 25000  # 100020 taken from chart_example.py  width in EMUs

def _generate_comparison_charts(lookup_df, col_indices, wb_obj, chart_sheet, nrow_chart, metric, max_sheet_row):
    """
    generate charts for each subarea for two sets of metrics
    """
    group_chart = LineChart()
    group_chart.style = 13

    defn, units, out_format, pyora_disp = fetch_detail_from_varname(lookup_df, metric)
    group_chart.title = defn
    group_chart.y_axis.title = units
    group_chart.x_axis.title = 'Time step'
    group_chart.height = 10
    group_chart.width = 60

    for iarea, col_indx in enumerate(col_indices):

        #
        # ======================================================
        sheet = wb_obj[metric]
        data = Reference(sheet, min_col=col_indx, min_row=2, max_col=col_indx, max_row=max_sheet_row)
        group_chart.add_data(data)  # title = subarea

        # Style the line just created
        # ===========================
        sref = group_chart.series[iarea]
        sref.graphicalProperties.line.width = PREFERRED_LINE_WIDTH

        dum, indx_color = divmod(iarea, 5)  # prevent IndexError by using remainder
        color_key = LINE_COLORS_SUBAREAS[indx_color]
        sref.graphicalProperties.line.solidFill = COLORS_DISTING[color_key]
        sref.smooth = True

    # now write to previously created sheet
    # =====================================
    chart_sheet.add_chart(group_chart, "B" + str(nrow_chart))
    nrow_chart += 20

    return nrow_chart

def _generate_pool_charts(sub_system, lookup_df, wb_obj, chart_sheet, nrow_chart):
    """
    initially for carbon only: generate charts for each subarea for two sets of metrics
    """
    for sht_indx, shtnm in enumerate(wb_obj.sheetnames):
        if shtnm not in ALPHABET:
            continue

        subarea = shtnm

        # construct header record
        # =======================
        sheet = wb_obj[subarea]
        max_sheet_row = sheet
        max_column = sheet.max_column - 1
        hdrs = [sheet[ch + '1'].value for ch in ALPHABET[:max_column]]

        for group_name in POOL_GROUPS:
            metrics_group = POOL_GROUPS[group_name]
            group_chart = LineChart()
            group_chart.style = 13
            if group_name == 'Active Pools':
                group_chart.title = subarea + '\t' + group_name
            else:
                group_chart.title = group_name

            group_chart.y_axis.title = 'Carbon (t/ha)'
            group_chart.x_axis.title = 'Time step'
            group_chart.height = 10
            group_chart.width = 20

            for isqnce, metric in enumerate(metrics_group):
                defn, units, out_format, pyora_disp = fetch_detail_from_varname(lookup_df, metric)

                if pyora_disp not in hdrs:
                    print(WARN_STR + pyora_disp + ' not in headers')
                    continue

                col_indx = hdrs.index(pyora_disp)
                data = Reference(sheet, min_col=col_indx, min_row=1, max_col=col_indx, max_row=max_sheet_row)
                group_chart.add_data(data)

                # Style the line just created
                # ===========================
                sref = group_chart.series[isqnce]
                sref.graphicalProperties.line.width = PREFERRED_LINE_WIDTH
                sref.graphicalProperties.line.solidFill = LINE_COLORS_POOL_GROUPS[isqnce]
                sref.smooth = True

            # now write to previously created sheet
            # =====================================
            if group_name == 'Active Pools':
                chart_sheet.add_chart(group_chart, "B" + str(nrow_chart))
            else:
                chart_sheet.add_chart(group_chart, "O" + str(nrow_chart))
                nrow_chart += 20

    return nrow_chart

def _generate_water_charts(col_indices, wb_obj, chart_sheet, nrow_chart, max_sheet_row):
    """
    water only: generate charts for each subarea
    """
    for col_indx in col_indices:
        subarea = col_indices[col_indx]
        if subarea is None:
            break
        metrics_group = list(['wat_soil', 'wat_drain'])
        group_chart = LineChart()
        group_chart.style = 13
        group_chart.title = subarea + '\tSteady state and forward run'
        group_chart.y_axis.title = ' Soil water content (mm)'
        group_chart.x_axis.title = 'Time step'
        group_chart.height = 10
        group_chart.width = 20

        # by design the sheet name is synonymous with the metric
        # ======================================================
        for isht, metric in enumerate(metrics_group):
            sheet = wb_obj[metric]
            data = Reference(sheet, min_col=col_indx, min_row=2, max_col=col_indx, max_row=max_sheet_row)
            group_chart.add_data(data)  # title = metric

            # Style the line just created
            # ===========================
            sref = group_chart.series[isht]
            sref.graphicalProperties.line.width = PREFERRED_LINE_WIDTH
            sref.graphicalProperties.line.solidFill = LINE_COLORS_POOL_GROUPS[isht]
            sref.smooth = True

        # write to previously created sheet
        # =================================
        chart_sheet.add_chart(group_chart, "B" + str(nrow_chart))
        nrow_chart += 20

    return nrow_chart

def _generate_charts(fname, sub_system, lookup_df):
    """
    add charts to an existing Excel file
    """
    # load workbook previously created
    # ================================
    if not exists(fname):
        print('File ' + fname + ' must exist')
        return -1

    wb_obj = load_workbook(fname, data_only=True)
    chart_sheet = wb_obj.create_sheet('charts')
    nrow_chart = 4  # locates top of chart

    if sub_system == 'carbon':
        nrow_chart = _generate_pool_charts(sub_system, lookup_df, wb_obj, chart_sheet, nrow_chart)
        # nrow_chart = _generate_comparison_charts(lookup_df, wb_obj, chart_sheet, nrow_chart, 'co2_emiss')

    elif sub_system == 'nitrogen':
        for metric in CHANGE_VARS['nitrogen']:
            # nrow_chart = _generate_comparison_charts(lookup_df, wb_obj, chart_sheet, nrow_chart, metric)
            pass

    elif sub_system == 'water':
        nrow_chart = _generate_water_charts(wb_obj, chart_sheet, nrow_chart)
        for metric in CHANGE_VARS['water']:
            # nrow_chart = _generate_comparison_charts(lookup_df, wb_obj, chart_sheet, nrow_chart, metric)
            pass
    try:
        wb_obj.active = len(wb_obj.sheetnames) - 1  # make the charts sheet active
        wb_obj.save(fname)
        print('\tcreated: ' + fname)
    except PermissionError as err:
        print(str(err) + ' - could not create: ' + fname)

    return

def write_excel_all_subareas(study, out_dir, lookup_df, all_runs):
    """
    Entry
    """
    for indx, sub_system in enumerate(CHANGE_VARS):

        # make a safe name
        # ===============
        sht_abbrev = sub_system
        fname = join(out_dir, study.study_name + ' z_' + sht_abbrev + '.xlsx')  # add 'z' to force order
        if isfile(fname):
            try:
                remove(fname)
            except PermissionError as err:
                print(err)
                return -1

        writer = ExcelWriter(fname)

        for subarea in all_runs:
            plot_dict = {'month': all_runs[subarea][0].data['imnth']}

            for metric in CHANGE_VARS[sub_system]:
                defn, units, out_format, pyora_disp = fetch_detail_from_varname(lookup_df, metric)
                plot_dict[pyora_disp] = all_runs[subarea][indx].data[metric]

            try:
                data_frame = DataFrame.from_dict(plot_dict)
            except ValueError as err:
                print(WARN_STR + str(err) + ' for variable ' + metric)
                data_frame = DataFrame()

            data_frame.to_excel(writer, subarea, index=False)

        writer.close()

        # reopen Excel file and make legible
        # ==================================
        MAX_WDTH = 15
        wb_obj = load_workbook(fname, data_only=True)
        for sheet_name in wb_obj.sheetnames:
            sheet = wb_obj[sheet_name]

            # column width
            # ============
            for ch in ALPHABET[1:sheet.max_column + 1]:
                sheet.column_dimensions[ch].width = MAX_WDTH
                cell_ref = ch + '1'
                sheet[cell_ref].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # rows
            # ====
            for irow in range(sheet.max_row + 1):
                sheet.row_dimensions[irow].height = 18
            sheet.row_dimensions[1].height = 48

        wb_obj.save(fname)
        print('\tadjusted row heights: ' + fname)

        # reopen Excel file and write charts
        # ==================================
        _generate_charts(fname, sub_system, lookup_df)
        print('\tadded charts to: ' + fname)

    return 0
