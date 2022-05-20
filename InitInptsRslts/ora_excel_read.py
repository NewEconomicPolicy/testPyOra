#-------------------------------------------------------------------------------
# Name:        ora_excel_read.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Definitions:
#   spin_up
#
# Description:
#
#
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'ora_excel_read.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
from os.path import isfile, isdir, split, normpath, join
from os import mkdir, sep as os_sep

from copy import copy
from openpyxl import load_workbook
from pandas import Series, read_excel, DataFrame
from zipfile import BadZipFile
from glob import glob
from calendar import monthrange
from pandas import DataFrame
from numpy import nan, isnan, array

from ora_water_model import add_pet_to_weather
from ora_cn_fns import plant_inputs_crops_distribution
from ora_low_level_fns import average_weather
from ora_classes_excel_write import pyoraId as oraId
from ora_gui_misc_fns import format_sbas, farming_system, region_validate, LivestockEntity

METRIC_LIST = list(['precip', 'tair'])
MNTH_NAMES_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

ANML_ABBREVS = ['catdry', 'catmt', 'rumdry', 'rummt', 'pigs', 'pltry']
ANML_PRODN_SHEET = 'Typical animal production'   # Herrero table

REQUIRED_SHEET_NAMES = list(['N constants', 'Crop parms', 'Org Waste parms', ANML_PRODN_SHEET])
FARM_WTHR_SHEET_NAMES = {'lctn': 'Farm location', 'wthr':'Weather'}

RUN_SHT_NAMES = {'sign': 'Signature', 'lctn': 'Farm location', 'wthr': 'Weather', 'sbas': 'Subareas',
                                                                                            'lvstck':'Livestock'}
SOIL_METRICS = list(['t_depth', 't_clay', 't_sand', 't_silt', 't_carbon', 't_bulk', 't_pH', 't_salinity'])
MNGMNT_SHT_HDRS = ['period', 'year', 'month', 'crop_name', 'yld', 'fert_type', 'fert_n', 'ow_type', 'ow_amnt', 'irrig']
T_DEPTH = 30

FNAME_RUN = 'FarmWthrMgmt.xlsx'

ERR_STR = '*** Error *** '
ERR_STR_SHEET = ERR_STR + 'reading sheet\t'
WARN_STR = '*** Warning *** '

from string import ascii_uppercase
ALPHABET = list(ascii_uppercase)
MAX_SUB_AREAS = 8
NFEED_TYPES = 5

class ReadLivestockSheet(object, ):

    def __init__(self, w_run_dir3, anml_prodn_obj):
        '''

        '''
        mgmt_dir = w_run_dir3.text()
        run_xls_fn = join(mgmt_dir, FNAME_RUN)

        wb_obj = load_workbook(run_xls_fn, data_only=True)

        print('Reading livestock sheet ')
        anml_abbrevs = ANML_ABBREVS

        lvstck_sht = wb_obj[RUN_SHT_NAMES['lvstck']]
        max_row = lvstck_sht.max_row
        max_col = lvstck_sht.max_column

        df = DataFrame(lvstck_sht.values, columns=['descr'] + anml_abbrevs)
        lvstck_dscrs = list(df.values[0][1:])

        wb_obj.close()

        # construct previous JSON derived dictionary
        lvstck_content = {'site definition': {'area name': 'all', 'area (ha)': 3.5, 'region': 'Eastern Africa',
                                                                                                    'system': 'MRA'}}
        # step through each animal type
        # =============================
        lvstck_indx = 0
        for anml, descr in zip(anml_abbrevs, lvstck_dscrs):
            nanmls = df[anml].values[1]
            if nanmls == 0:
                continue
            strtgy =  df[anml].values[2]
            lvstck = {'type': descr, 'number': nanmls, 'strategy': strtgy}

            indx = 3
            feed_indx = 0
            for ic in range(NFEED_TYPES):
                feed_type = df[anml].values[indx]
                val = df[anml].values[indx + 1]
                if feed_type != 'None' and val > 0:
                    feed_indx += 1
                    feed_id = 'feed' + str(feed_indx)
                    lvstck[feed_id] = {'type': feed_type, 'value': val}

                indx += 2

            # check for bought in
            # ===================
            val = df[anml].values[-1]
            if val is not None:
                feed_indx += 1
                feed_id = 'feed' + str(feed_indx)
                lvstck[feed_id] = {'type': 'bought in', 'value': float(val)}


            if feed_indx > 0:
                lvstck_indx += 1
                lvstck_id = 'livestock' + str(lvstck_indx)
                lvstck_content['site definition'][lvstck_id] = lvstck

        # from JSON legacy
        # ================
        site_defn = lvstck_content['site definition']
        area = site_defn['area name']

        region = region_validate(site_defn, anml_prodn_obj)
        system = farming_system(site_defn)

        lvstck_grp = []
        for key in site_defn:
            if key.find('livestock') > -1:
                lvstck_grp.append(LivestockEntity(site_defn[key], anml_prodn_obj))

        subareas = {}
        subareas[area] = {'region': region, 'system': system, 'lvstck_grp': lvstck_grp}

        self.subareas = subareas

        print()  # cosmetic
        '''
        for anml, nmbr in zip(anml_abbrevs, df.values[irow][1:]):  # numbers of animals
            if nmbr is not None:
                pass
        irow += 1
        for anml, strtg in zip(anml_abbrevs, df.values[irow][1:]):  # strategies
            pass    # strtg

        for findx in range(NFEED_TYPES):
            irow += 1
            fd_typ = str(findx + 1)
            for anml, feed_typ in zip(anml_abbrevs, df.values[irow][1:]):  # feed type
                pass  # feed_typ

            irow += 1
            for anml, feed_qty in zip(anml_abbrevs, df.values[irow][1:]):  # feed quantity as a percentage
                pass  # feed_qty

        irow += 1
        for anml, pcnt in zip(anml_abbrevs, df.values[irow][1:]):  # percentage bought in
            if pcnt is not None:
                pass  # pcnt     
        '''

def _create_ow_fert(df):
    '''
'   create fertiliser and organic waste lists from series
    '''
    fert_ns = []
    for fert_type, fert_n in zip(df['fert_type'].values, df['fert_n'].values):
        if fert_type is None:
            fert_ns.append(None)
        else:
            fert_ns.append({'fert_type': fert_type, 'fert_n': fert_n})

    org_ferts = []
    for ow_type, ow_amnt in zip(df['ow_type'].values, df['ow_amnt'].values):
        if ow_type is None:
            org_ferts.append(None)
        else:
            org_ferts.append({'ow_type': ow_type, 'amount': ow_amnt})

    irrigs_tmp = list(df['irrig'].values)
    irrigs = []
    for irrig in irrigs_tmp:
        if irrig is None:
            irrigs.append(0)
        else:
            if isnan(irrig):
                irrigs.append(0)
            else:
                irrigs.append(irrig)

    return fert_ns, org_ferts, irrigs

def _add_tgdd_to_weather(tair_list):
    '''
    growing degree days indicates the cumulative temperature when plant growth is assumed to be possible (above 5Â°C)
    '''
    imnth = 1
    grow_dds = []
    for tair in tair_list:

        dummy, ndays = monthrange(2011, imnth)
        n_grow_days = max(0, ndays * (tair - 5))    #  (eq.3.2.2)
        grow_dds.append(round(n_grow_days,3))

        imnth += 1
        if imnth > 12:
            imnth = 1

    return grow_dds

def _validate_timesteps(run_xls_fn, subareas):
    '''
    for each subarea, check number of months against weather
    '''
    ret_code = False
    dum, farm_name = split(split(run_xls_fn)[0])
    mess = 'Farm: ' + farm_name + '\t'

    wb_obj = load_workbook(run_xls_fn, data_only=True)

    # check for Nones in weather
    # ==========================
    wthr_sht = wb_obj[RUN_SHT_NAMES['wthr']]
    wthr_cols = ['period', 'year', 'month', 'precip', 'tair']
    if wthr_sht.max_column == 6:
        wthr_cols += ['actl_yr']
    try:
        df = DataFrame(wthr_sht.values, columns=wthr_cols)

        '''
        # check for None - TODO: seems unreliable
        # ==============
        mess_null = ''
        if not df.isnull().values.any():
            mess += '\tNulls encountered in weather sheet'
            ret_code = False
        '''

    except ValueError as err:
        print(ERR_STR + 'reading run file weather sheet: ' + str(err))
        return ret_code

    # identify issues to warn user
    # ============================
    ret_code = True
    warn_mess = ''
    nmnths_wthr = wthr_sht.max_row - 1
    mess += 'weather months {} '.format(nmnths_wthr)

    # check subareas
    # ==============
    nmnths_subareas = {}
    for sba in subareas:
        sba_sht = wb_obj[sba]
        nmnths_sba = sba_sht.max_row - 1
        nmnths_subareas[sba] = nmnths_sba

    sba_mnths = array(list(nmnths_subareas.values()))
    if all(sba_mnths == sba_mnths[0]):
        mess += '\tsubarea months: {}'.format(sba_mnths[0])
        if sba_mnths[0] != nmnths_wthr:
            warn_mess += 'different subarea and weather months'
    else:
        warn_mess += 'subarea sheets have inconsistent number of months: ' + str(nmnths_subareas)

    wb_obj.close

    # add warnings
    # ============
    if len(warn_mess) > 0:
        mess += ' ' + WARN_STR + warn_mess

    print(mess)

    return ret_code

def check_xls_run_file(w_run_model, mgmt_dir):
    '''
    =========== called during initialisation or from GUI when changing farm ==============
    validate xls run file
    '''
    w_run_model.setEnabled(False)
    farm_wthr_fname = FNAME_RUN
    mess = 'Run file, ' + farm_wthr_fname + ', is '

    run_xls_fn = join(mgmt_dir, farm_wthr_fname)
    if not isfile(run_xls_fn):
        mess += 'non existent'
        return mess

    # check required sheets are present
    # =================================
    integrity_flag = True
    wb_obj = load_workbook(run_xls_fn, data_only=True)
    for rqrd_sheet in RUN_SHT_NAMES.values():
        if rqrd_sheet not in wb_obj.sheetnames:
            print('Sheet ' + rqrd_sheet + ' not present in ' + run_xls_fn)
            integrity_flag = False
    wb_obj.close()

    if not integrity_flag:
        mess += 'uncompliant'
        return mess

    ret_var = read_farm_wthr_xls_file(run_xls_fn)
    if ret_var is None:
        mess += 'uncompliant'
    else:
        subareas = ret_var[0]
        mess = format_sbas(subareas)
        if (_validate_timesteps(run_xls_fn, subareas)):
            w_run_model.setEnabled(True)      # activate carbon nitrogen model push button

    return mess

def read_xls_run_file(run_xls_fn, crop_vars, latitude):
    '''
    check required sheets are present and read data from these
     '''
    ret_var = None
    wb_obj = load_workbook(run_xls_fn, data_only=True)

    # check required sheets are present
    # =================================
    for rqrd_sheet in RUN_SHT_NAMES.values():
        if rqrd_sheet not in wb_obj.sheetnames:
            print('Sheet ' + rqrd_sheet + ' not present in ' + run_xls_fn)
            return ret_var

    # Farm location - not used
    # =============
    lctn_sht = wb_obj[RUN_SHT_NAMES['lctn']]
    df = DataFrame(lctn_sht.values, columns=['Attribute', 'Values'])
    lctn_var = list(df['Values'][:])

    # Subareas - read subareas lookup sheet which includes soil definition
    # ====================================================================
    sbas_sht = wb_obj[RUN_SHT_NAMES['sbas']]
    rows_generator = sbas_sht.values
    hdr_row = next(rows_generator)  # skip headers
    data_rows = [list(row) for (_, row) in zip(range(MAX_SUB_AREAS), rows_generator)]

    ora_subareas = {}
    for rec in data_rows:
        descr = rec[1]

        # if subarea description is present then make sure the corresponding subarea sheet is present
        # ===========================================================================================
        if descr is not None:
            sba = rec[0]
            if sba not in wb_obj:
                print(WARN_STR + 'Management sheet ' + sba + ' not in run file')
                continue

            area_ha = rec[4]

            # soils
            # =====
            soil_defn = {}
            for val, metric in zip([T_DEPTH] + rec[5:], SOIL_METRICS):
                if val is None:
                    print(ERR_STR + 'Soil for subarea ' + sba + ' not defined in run file')
                    return ret_var
                else:
                    soil_defn[metric] = val

            # read subarea sheet
            # ==================
            soil_for_area = Soil(soil_defn)
            ora_subareas[sba] = ReadMngmntSubareas(wb_obj, sba, soil_for_area, crop_vars, area_ha)

    # Weather sheet can have 5 or 6 columns
    # =====================================
    pettmp_ss = {'precip': [], 'tair': []}
    pettmp_fwd = {'precip': [], 'tair': []}

    wthr_sht = wb_obj[RUN_SHT_NAMES['wthr']]
    wthr_cols = ['period', 'year', 'month', 'precip', 'tair']
    if wthr_sht.max_column == 6:
        wthr_cols += ['actl_yr']
    try:
        df = DataFrame(wthr_sht.values, columns = wthr_cols)
    except ValueError as err:
        print(ERR_STR + 'reading run file weather sheet: ' + str(err))
        return ret_var

    iline = 1
    for mode, precip, tair in zip(df['period'].values[1:], df['precip'].values[1:], df['tair'].values[1:]):
        iline += 1
        if mode is None or precip is None or tair is None:
            print(ERR_STR + 'null values encountered on line {} when reading run file weather sheet'.format(iline))
            return ret_var

        if mode == 'steady state':
            pettmp_ss['precip'].append(precip)
            pettmp_ss['tair'].append(tair)
        else:
            pettmp_fwd['precip'].append(precip)
            pettmp_fwd['tair'].append(tair)

    # correct any temporal misalignment between weather and management
    # ================================================================
    ntsteps_fwd = ora_subareas[sba].ntsteps_fwd
    ntsteps_wthr = len(pettmp_fwd['precip'])
    if ntsteps_wthr != ntsteps_fwd:
        pettmp_fwd = _sync_wthr_to_mgmt(pettmp_fwd, ntsteps_wthr, ntsteps_fwd)

    ora_weather = WeatherRelated(pettmp_ss, pettmp_fwd, latitude)  # construct weather object including degree days etc

    wb_obj.close()

    # wrap up
    # =======
    if len(ora_subareas) == 0:
        print(ERR_STR + 'the Subareas sheet must have at least one subarea with a description - check run file')
    else:
        ret_var = (ora_weather, ora_subareas)

    return ret_var

def _sync_wthr_to_mgmt(pettmp, ntsteps_wthr, ntsteps_sim):
    '''
    Truncate or stretch supplied weather to match length for simulation period
    '''
    if ntsteps_sim > ntsteps_wthr:
        mess = 'Stretched'
    else:
        mess = 'Truncated'

    nsim_yrs = int(ntsteps_sim / 12)
    nwthr_yrs = int(ntsteps_wthr / 12)
    print(mess + ' weather from {} years to match simulation period of {} years'.format(nwthr_yrs, nsim_yrs))

    NPERIODS = 10
    pettmp_sim = {}
    for metric in pettmp:
        pettmp_lst = []
        for iyr in range(NPERIODS):
            pettmp_lst += pettmp[metric]
            if len(pettmp_lst) > ntsteps_sim:
                pettmp_sim[metric] = pettmp_lst[:ntsteps_sim]
                break

    return pettmp_sim

def _make_current_crop_list(crop_names):
    '''
    fill in crops for entire period
    '''
    crop_currs = copy(crop_names)
    prev_crop = crop_names[0]
    for indx, crop in enumerate(crop_names):
        if crop is None:

            # if remaining entries are all None then fill with previous crop
            # ==============================================================
            rmn_set = set(crop_names[indx:])
            if len(rmn_set) == 1:
                if list(rmn_set)[0] is None:
                    crop_currs[indx] = prev_crop

            # find next crop in future months
            # ===============================
            for crop_fwd in crop_names[indx + 1:]:
                if crop_fwd is not None:
                    crop_currs[indx] = crop_fwd
                    break
        else:
            # remember last crop
            # ==================
            prev_crop = crop

    return crop_currs

def _amend_pi_props_tonnes(crop_vars, this_crop, indx_strt, pi_props, pi_tonnes):
    '''

    '''
    crop = this_crop[0]
    ngrow_mnths = len(this_crop)
    indx_end = indx_strt + ngrow_mnths

    if ngrow_mnths == crop_vars[crop]['t_grow']:
        pi_tonnes[indx_strt:indx_end] = crop_vars[crop]['pi_tonnes']
        pi_props[indx_strt:indx_end] = crop_vars[crop]['pi_prop']
    else:
        pi_tonnes[indx_strt:indx_end] = ngrow_mnths*[999]
        pi_props[indx_strt:indx_end] = ngrow_mnths*[999]

    yield_typ = crop_vars[crop]['max_yld']  # TODO: should get yield from run file

    return Crop(crop, yield_typ)

def _make_pi_props_tonnes(crop_names, indx_mode, crop_vars):
    '''
    accumulate growing months for each crop
    consider contiguous perennial crops - grassland, scrubland, coffee
    '''
    func_name = __prog__ + '\t_make_pi_props_tonnes'

    ntsteps = len(crop_names)
    pi_props = ntsteps*[0]
    pi_tonnes = ntsteps*[0]

    crops_ss = []
    crops_fwd = []

    indx_strt = None
    prev_crop = None
    this_crop = []

    for indx, crop in enumerate(crop_names):
        if crop is not None:
            if indx_strt is None:
                indx_strt = indx
                this_crop = [crop]

            elif crop == prev_crop:
                this_crop.append(crop)

            # elif prev_crop is None and len(this_crop) == 1:
            #     this_crop = [crop]
            else:
                if len(this_crop) > 0:
                    # record plant inputs for this crop
                    # =================================
                    crop_obj = _amend_pi_props_tonnes(crop_vars, this_crop, indx_strt, pi_props, pi_tonnes)
                    if indx_strt < indx_mode:
                        crops_ss.append(crop_obj)
                    else:
                        crops_fwd.append(crop_obj)

                    this_crop = [crop]
                    indx_strt = indx

        prev_crop = crop

    if len(this_crop) > 0:
        # check if final crop remains
        # ===========================
        crop_obj = _amend_pi_props_tonnes(crop_vars, this_crop, indx_strt, pi_props, pi_tonnes)
        crops_fwd.append(crop_obj)

    return pi_props, pi_tonnes, crops_ss, crops_fwd

class Crop(object,):
    '''

    '''
    def __init__(self, crop_name, yield_typ):
        """
        Assumptions:
        """
        self.crop_lu = crop_name
        self.yield_typ = yield_typ

class ReadMngmntSubareas(object, ):

    def __init__(self, wb_obj, sba, soil_for_area, crop_vars, area_ha):
        '''

        '''
        print('Reading management sheet ' + sba)

        mgmt_sht = wb_obj[sba]
        ntsteps = mgmt_sht.max_row - 1
        rows_generator = mgmt_sht.values
        header_row = next(rows_generator)
        data_rows = [list(row) for (_, row) in zip(range(ntsteps), rows_generator)]

        df = DataFrame(data_rows, columns = MNGMNT_SHT_HDRS)
        period_list = list(df['period'].values)
        try:
            indx_mode = period_list.index('forward run')
        except ValueError as err:
            print(ERR_STR + 'bad subarea sheet ' + sba)
            return

        crop_names = list(df['crop_name'].values)

        crop_currs = _make_current_crop_list(crop_names)
        fert_n_list, org_fert_list, irrigs = _create_ow_fert(df)
        pi_props, pi_tonnes, crops_ss, crops_fwd = _make_pi_props_tonnes(crop_names, indx_mode, crop_vars)

        # TODO: crude and unpythonic
        # ==========================
        crop_mngmnt_fwd = {}
        crop_mngmnt_ss = {}

        crop_mngmnt_ss['crop_name'] = crop_names[:indx_mode]
        crop_mngmnt_ss['crop_curr'] = crop_currs[:indx_mode]
        crop_mngmnt_ss['crop_mngmnt'] = crops_ss
        crop_mngmnt_ss['fert_n'] = fert_n_list[:indx_mode]
        crop_mngmnt_ss['org_fert'] = org_fert_list[:indx_mode]
        crop_mngmnt_ss['pi_prop'] = pi_props[:indx_mode]
        crop_mngmnt_ss['pi_tonne'] = pi_tonnes[:indx_mode]
        crop_mngmnt_ss['irrig'] = irrigs[:indx_mode]

        crop_mngmnt_fwd['crop_name'] = crop_names[indx_mode:]
        crop_mngmnt_fwd['crop_curr'] = crop_currs[indx_mode:]
        crop_mngmnt_fwd['crop_mngmnt'] = crops_fwd
        crop_mngmnt_fwd['fert_n'] = fert_n_list[indx_mode:]
        crop_mngmnt_fwd['org_fert'] = org_fert_list[indx_mode:]
        crop_mngmnt_fwd['pi_prop'] = pi_props[indx_mode:]
        crop_mngmnt_fwd['pi_tonne'] = pi_tonnes[indx_mode:]
        crop_mngmnt_fwd['irrig'] = irrigs[indx_mode:]

        self.soil_for_area = soil_for_area
        self.crop_mngmnt_ss = crop_mngmnt_ss

        self.crop_mngmnt_fwd = crop_mngmnt_fwd
        self.area_ha = area_ha
        self.ntsteps_ss = len(crop_mngmnt_ss['crop_name'])
        self.ntsteps_fwd = len(crop_mngmnt_fwd['crop_name'])

class WeatherRelated(object, ):

    def __init__(self, pettmp_ss, pettmp_fwd, latitude):
        '''
        onstruct weather object including degree days
        '''
        # generate PET from weather
        # =========================
        self.pettmp_ss = add_pet_to_weather(latitude, pettmp_ss)
        self.pettmp_fwd = add_pet_to_weather(latitude, pettmp_fwd)

        # growing degree days
        # ==================
        self.pettmp_ss['grow_dds'] = _add_tgdd_to_weather(pettmp_ss['tair'])
        self.pettmp_fwd['grow_dds'] = _add_tgdd_to_weather(pettmp_fwd['tair'])

        # average monthly precip and temp is required for spin up
        # =======================================================
        self.ave_precip_ss, self.ave_temp_ss, self.ave_pet_ss = \
                            average_weather(latitude, self.pettmp_ss['precip'], self.pettmp_ss['tair'])

        # get average annual rain and temperature of first 10 years
        # =========================================================
        nmnths = len(pettmp_ss['precip'])
        nyrs = nmnths/12
        self.ann_ave_precip_ss = sum(pettmp_ss['precip'])/nyrs
        self.ann_ave_temp_ss = sum(pettmp_ss['tair'])/nmnths

class Soil(object,):
    '''

    '''
    def __init__(self, soil_defn):
        """
        Assumptions:
        """
        self.title = 'Soil'

        t_depth = soil_defn['t_depth']
        self.t_clay = soil_defn['t_clay']
        self.t_silt = soil_defn['t_silt']
        self.t_sand = soil_defn['t_sand']
        t_carbon = soil_defn['t_carbon']
        t_bulk = soil_defn['t_bulk']
        self.t_pH_h2o = soil_defn['t_pH']

        salinity = soil_defn['t_salinity']
        if salinity is None:
            self.t_salinity = 0.0
        else:
            self.t_salinity = salinity

        tot_soc_meas = (10 ** 4) * (t_depth / 100) * t_bulk * (t_carbon / 100)  # tonnes/ha

        self.t_depth = t_depth
        self.t_carbon = t_carbon
        self.t_bulk = t_bulk
        self.tot_soc_meas = tot_soc_meas

def check_params_excel_file(params_xls_fn):
    '''
    validate selected Excel parameters file and disable CN model push button if not valid
    '''
    retcode = None

    if not isfile(params_xls_fn):
        print(ERR_STR + 'Excel parameters file ' + params_xls_fn + ' must exist - check setup file')
        return None

    print('ORATOR parameters file: ' + params_xls_fn)
    try:
        wb_obj = load_workbook(params_xls_fn, data_only = True)
        sheet_names = wb_obj.sheetnames
    except (PermissionError, BadZipFile) as err:
        print(ERR_STR + str(err))
        return retcode

    wb_obj.close()

    # all required sheets must be present
    # ===================================
    for sheet in REQUIRED_SHEET_NAMES:
        if sheet not in sheet_names:
            print(ERR_STR + 'Required sheet ' + sheet + ' is not present - please check file')
            return retcode

    return 0

def _read_n_constants_sheet(xls_fname, sheet_name, skip_until):
    '''
    r_dry is an environmental constant
    '''

    n_parm_names = list(['atmos_n_depos', 'prop_atmos_dep_no3', 'no3_min', 'k_nitrif',
                      'n_denit_max', 'n_d50', 'prop_n2o_fc', 'prop_nitrif_gas', 'prop_nitrif_no',
                      'precip_critic', 'prop_volat', 'prop_atmos_dep_nh4', 'c_n_rat_soil', 'r_dry', 'k_c_rate'])

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(1,3))
    n_parms_df = data.dropna(how='all')
    n_parms = {}
    for indx, defn in enumerate(n_parm_names):
        n_parms[defn]  = n_parms_df['Value'].values[indx]

    return n_parms

def _read_crop_vars(xls_fname, sheet_name):
    '''
    read maximum rooting depths etc. for each crop from sheet A1c
    '''
    crop_parm_names = Series(list(['lu_code', 'rat_dpm_rpm', 'harv_indx', 'prop_npp_to_pi', 'max_root_dpth',
                    'max_root_dpth_orig', 'sow_mnth', 'harv_mnth', 'max_yld', 'dummy1',
                    'c_n_rat_pi', 'n_sply_min', 'n_sply_opt', 'n_respns_coef', 'fert_use_eff',
                    'dummy3', 'n_rcoef_a', 'n_rcoef_b', 'n_rcoef_c', 'n_rcoef_d', 'gdds_scle_factr','iws_scle_factr']))

    data = read_excel(xls_fname, sheet_name)
    data = data.dropna(how='all')
    try:
        crop_dframe = data.set_index(crop_parm_names)
        crop_vars = crop_dframe.to_dict()
    except ValueError as err:

        print(ERR_STR_SHEET + sheet_name + ' ' + str(err))
        return None

    # discard unwanted entries
    # ========================
    for crop in ['Crop', 'None', 'Null']:
        del(crop_vars[crop])

    for crop_name in crop_vars:

        # clean data
        # ==========
        for var in ['harv_mnth', 'sow_mnth', 'lu_code']:
            crop_vars[crop_name][var] = int(crop_vars[crop_name][var])

        # number of months in the growing season
        # ======================================
        harv_mnth = crop_vars[crop_name]['harv_mnth']
        sow_mnth = crop_vars[crop_name]['sow_mnth']
        if sow_mnth > harv_mnth:
            harv_mnth += 12
        t_grow = harv_mnth - sow_mnth + 1
        crop_vars[crop_name]['pi_tonnes'], crop_vars[crop_name]['pi_prop'] = plant_inputs_crops_distribution(t_grow)
        crop_vars[crop_name]['t_grow'] = t_grow

    return crop_vars

def _read_organic_waste_sheet(xls_fname, sheet_name, skip_until):
    '''
    read Organic waste parameters
    added  - see (eq.2.1.12) and (eq.2.1.13)
    TODO percentages are converted to fraction
    '''
    ow_parms_names = Series(list(['c_n_rat', 'prop_nh4', 'rat_dpm_hum_ow', 'prop_iom_ow', 'pcnt_c', 'min_e_pcnt_wd',
                                  'max_e_pcnt_wd', 'ann_c_input', 'pcnt_urea']))

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until))
    data = data.dropna(how='all')
    try:
        ow_dframe = data.set_index(ow_parms_names)
        all_ow_parms = ow_dframe.to_dict()
    except ValueError as err:
        print(ERR_STR_SHEET + sheet_name + ' ' + str(err))
        all_ow_parms = None

    return all_ow_parms

def read_econ_purch_sales_sheet(xls_fname, sheet_name, skip_until):
    '''
    Read data on purchases and sales, required for econ module
    '''

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until))
    purch_sales_df = DataFrame(data)

    return purch_sales_df

def read_econ_labour_sheet(xls_fname, sheet_name, skip_until):
    '''
    Read data on labour, required for econ module
    '''

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until))
    labour_df = DataFrame(data)

    return labour_df

def _repopulate_excel_dropdown(form, study_name):
    '''
    repopulate Excel drop-down associated with Display output Excel files
    '''
    if hasattr(form, 'w_combo17'):
        w_combo17 = form.w_combo17
        w_disp_out = form.w_disp_out
    else:
        w_combo17 = form.w_tab_wdgt.w_combo17
        w_disp_out = form.w_tab_wdgt.w_disp_out

    out_dir = form.settings['out_dir']
    xlsx_list = glob(out_dir + '/' + study_name + '*.xlsx')
    w_combo17.clear()
    if len(xlsx_list) > 0:
        w_disp_out.setEnabled(True)
        for out_xlsx in xlsx_list:
            dummy, short_fn = split(out_xlsx)
            w_combo17.addItem(short_fn)
    return

def _make_retvar_safe(ret_var):
    '''
    ret_var is a list
    make sure return value is safe
    '''
    NVALS_SAFE = 10

    nvals = len(ret_var)
    if nvals > NVALS_SAFE:
        ret_var = ret_var[:NVALS_SAFE]
    elif nvals < NVALS_SAFE:
        nvals_add = NVALS_SAFE - nvals
        ret_var += 2*[None]

    return ret_var

def read_farm_wthr_xls_file(run_xls_fn):
    '''
    check required sheets are present
    '''
    wb_obj = load_workbook(run_xls_fn, data_only=True)

    subareas = []
    for sht in wb_obj.sheetnames:
        if sht in oraId().SUBAREAS:
            subareas.append(sht)

    subareas.sort()     # returns null value
    ret_var = list([subareas])

    rqrd_sheet = FARM_WTHR_SHEET_NAMES['lctn']
    if rqrd_sheet in wb_obj.sheetnames:
        farm_sht = wb_obj[rqrd_sheet]
        df = DataFrame(farm_sht.values, columns=['Attribute', 'Values'])
        ret_var += list(df['Values'][1:])
        ret_var = _make_retvar_safe(ret_var)
    else:
        print('Sheet ' + rqrd_sheet + ' not present in ' + run_xls_fn)
        ret_var = None

    wb_obj.close()

    return ret_var

class ReadStudy(object, ):

    def __init__(self, form, mgmt_dir, run_xls_fname = None, output_excel = True):
        '''
        read location sheet from ORATOR inputs Excel file
        '''
        self.output_excel = output_excel

        if run_xls_fname is None:
            run_xls_fname = normpath(join(mgmt_dir, FNAME_RUN))

        if not isfile(run_xls_fname):
            print('No run file ' + FNAME_RUN + ' in directory ' + mgmt_dir)
            return None

        # split management directory
        # ==========================
        norm_path = normpath(mgmt_dir)
        path_cmpnts = norm_path.split(os_sep)
        study_area, farm = path_cmpnts[-2:]

        # Farm location
        # =============
        ret_var = read_farm_wthr_xls_file(run_xls_fname)
        if ret_var is None:
            return None
        
        subareas, sub_distr, farm_name, latitude, longitude, area, prnct, dum, dum, dum = ret_var

        # check consistency
        # =================
        if farm_name != farm:
            print(WARN_STR + 'Inconsistent farm names: ' + farm + '\tname in run file: ' + farm_name)

        study_desc = 'Study area: ' + study_area + '\t\tFarm: ' + farm_name
        study_desc += '\t\tLatitude: {}'.format(latitude)
        study_desc += '\tLongitude: {}'.format(longitude)
        if hasattr(form, 'w_study'):
            form.w_study.setText(study_desc)
        else:
            form.w_tab_wdgt.w_study.setText(study_desc)

        self.study_name = farm_name
        self.latitude = latitude
        self.longitude = longitude
        self.subareas = subareas

        '''
         base outputs directory on inputs location, check and if necessary create
        '''
        out_dir = normpath(join(mgmt_dir, 'outputs'))
        if not isdir(out_dir):
            try:
                mkdir(out_dir)
                print('Created output directory: ' + out_dir)
            except:
                raise Exception('*** Error *** Could not create output directory: ' + out_dir)

        form.settings['out_dir'] = out_dir
        _repopulate_excel_dropdown(form, farm_name)

class ReadAnmlProdn(object, ):

    def __init__(self, xls_fname, crop_vars):
        '''
        read values from sheet C1a: Typical animal production in Africa provided Herrero et al. (2016)
        '''
        func_name =  __prog__ +  ' oratorExcelDetail __init__'

        self.retcode = None
        self.header_mappings = {'Type': 'Livestock type', 'ProdSystem': 'Livestock production system',
             'Region': 'Region', 'System': 'System', 'Milk': 'Milk', 'Meat': 'Meat',
             'FSgraze': 'Feedstock dry matter from grazing',
             'FSstovers': 'Feedstock dry matter from stovers',
             'FSoccas': 'Feedstock dry matter from occasional sources',
             'FSgrain': 'Feedstock dry matter from grain',
             'Manure': 'Manure dry matter', 'ExcreteN': 'Excreted N'}

        print('Reading animal production data from sheet: ' + ANML_PRODN_SHEET)
        column_names = 	list(self.header_mappings.keys())
        data = read_excel(xls_fname, header=None, names= column_names, sheet_name=ANML_PRODN_SHEET,
                                                                        usecols=range(1,13), skiprows=range(0,13))
        anml_prodn = data.dropna(how='all')
        self.anml_prodn = anml_prodn

        # allowable values required for validation
        # ========================================
        self.anml_types = list(anml_prodn['Type'].unique())   # + list(['Pigs','Poultry'])
        self.prodn_systms = list(anml_prodn['ProdSystem'].unique())
        self.world_regions = list(anml_prodn['Region'].unique())
        self.farm_systems = list(anml_prodn['System'].unique())
        self.crop_names = ['None'] + list(crop_vars.keys())

        # TODO: a patch to get through transition
        # create dictionary of generic animal types
        # =======================================
        anml_abbrevs = ['catdry','catmt','rumdry','rummt','pigs','pltry']
        gnrc_anml_types = {}
        for abbrev, anml_typ in zip(anml_abbrevs, self.anml_types):
            gnrc_anml_types[abbrev] = anml_typ
        self.gnrc_anml_types = gnrc_anml_types

        self.retcode = 0

def read_subarea_sheet(wthr_xls, sba_indx, nyrs_rota, mngmnt_hdrs):
    '''
    read first rotation period of management sheet
    '''
    df = None
    data_rows = None

    nmnths = 12*nyrs_rota
    wb_obj = load_workbook(wthr_xls, data_only=True)
    if sba_indx in wb_obj.sheetnames:
        sba_sht = wb_obj[sba_indx]
        rows_generator = sba_sht.values
        header_row = next(rows_generator)
        data_rows = [list(row) for (_, row) in zip(range(nmnths), rows_generator)]

        # df = DataFrame(data_rows, columns=mngmnt_hdrs)

    wb_obj.close()

    if data_rows is None:
        return data_rows

    if len(data_rows[0]) < 9:
        print(ERR_STR_SHEET + sba_indx + ' must have at least 9 values per line')
        return None

    # condition data so yield, Fert N amount, OW amount and	irrigation are mapped from None to zero
    # =============================================================================================
    data_recs = []
    for row in data_rows:
        new_row = copy(row)
        for icol in list([4, 5, 8, 9]):     # corresponds to columns E, G, I and J
            if row[icol] is None:
                new_row[icol] = '0'

        data_recs.append(new_row)

    return data_recs


class ReadCropOwNitrogenParms(object, ):

    def __init__(self, params_xls_fn):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading crop, organic waste and Nitrogen parameter sheets...')

        # Nitrogen params plus r_dry, drying potential
        # ============================================
        self.n_parms = _read_n_constants_sheet(params_xls_fn, 'N constants', 0)

        # Organic Waste and Crop params e.g. max rooting depths
        # =====================================================
        self.ow_parms = _read_organic_waste_sheet(params_xls_fn, 'Org Waste parms', 0)
        self.crop_vars = _read_crop_vars(params_xls_fn, 'Crop parms')
