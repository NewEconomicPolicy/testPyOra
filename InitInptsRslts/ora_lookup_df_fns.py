#-------------------------------------------------------------------------------
# Name:        ora_lookup_df_fns.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>#
# Description:
#   lookup table has followimg headers:
#       "PyOrator variable"	Category	"PyOrator display"	Symbol	Definition	Units	"Output format"	Notes
#-------------------------------------------------------------------------------

__prog__ = 'ora_lookup_df_fns.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
from os.path import isfile
from numpy import nan
from openpyxl import load_workbook
from pandas import read_excel
from zipfile import BadZipFile

from thornthwaite import thornthwaite

ERROR_STR = '*** Error *** '
WARN_STR = '*** Warning *** '

PY_VAR = 'PyOrator variable'
PY_DISP = 'PyOrator display'
APPNDX_A_SHEET = 'Appendix A'

def fetch_display_names_from_metrics(lookup_df, category_change):
    '''
    return list of PyOrator display names from lookup data frame for this category where
    category can be carbon, nitrogen or soil water
    '''
    metric_list = []
    for metric in category_change.data.keys():
        if metric == 'imnth' or metric == 'tstep' or metric == 'crop_name':
            continue
        else:
            metric_list.append(metric)

    display_names = []
    for metric in metric_list:
        result = lookup_df[PY_VAR][lookup_df[PY_VAR] == metric]
        if len(result) > 0:
            key = result.index[0]
            pyora_display = str(lookup_df[PY_DISP][key])
            if pyora_display == 'nan' or pyora_display is None:
                print(WARN_STR + ' skipping metric ' + metric +
                                                ' - could not retrieve PyOrator display name from lookup data frame')
                continue

            display_names.append(pyora_display)

    return sorted(display_names)

def fetch_detail_from_varname(lookup_df, metric):
    '''
    retrieve detail associated with metric if it is present
    '''
    dflt_rtrn = list([metric, '', '', metric])
    result = lookup_df[PY_VAR][lookup_df[PY_VAR] == metric]
    if len(result) == 0:
        return dflt_rtrn
    else:
        key = result.index[0]
        definition = lookup_df['Definition'][key]
        units = lookup_df['Units'][key]
        out_format = lookup_df['Output format'][key]
        if out_format is nan:
            out_format = '2f'

        pyora_display = lookup_df[PY_DISP][key]

        return definition, units, out_format, pyora_display

def fetch_display_from_varname(lookup_df, metric):
    '''
    retrieve variable display name from metric if it is present
    '''
    result = lookup_df[PY_VAR][lookup_df[PY_VAR] == metric]
    if len(result) == 0:
        return None
    else:
        key = result.index[0]
        pyora_display = lookup_df[PY_DISP][key]
        if pyora_display is nan:
            return None

        return pyora_display

def fetch_pyora_varname_from_pyora_display(lookup_df, pyora_display):
    '''
    return PyOrator variable name from data frame for PyOrator display value if found
    '''
    result = lookup_df[PY_DISP][lookup_df[PY_DISP] == pyora_display]
    if len(result) == 0:
        return None
    else:
        key = result.index[0]
        varname = lookup_df[PY_VAR][key]

        return varname

def fetch_definition_from_pyora_display(lookup_df, pyora_display):
    '''
    return symbol definition from data frame for PyOrator display value if found
    '''
    result = lookup_df[PY_DISP][lookup_df[PY_DISP] == pyora_display]
    if len(result) == 0:
        return list([pyora_display, ''])
    else:
        key = result.index[0]
        defn = lookup_df['Definition'][key]

        return defn

def fetch_defn_units_from_pyora_display(lookup_df, pyora_display):
    '''
    retrieve variable definition from data frame if it is present
    '''
    dflt_rtrn = list([pyora_display, ''])
    if lookup_df is None:
        return dflt_rtrn

    result = lookup_df[PY_VAR][lookup_df[PY_DISP] == pyora_display]
    if len(result) == 0:
        return dflt_rtrn
    else:
        key = result.index[0]
        defn = lookup_df['Definition'][key]
        units = lookup_df['Units'][key]
        if not isinstance(units, str):
            units = ''
        return defn, units

def read_lookup_excel_file(settings, batch_flag=False):
    '''
    check to see if a lookup Excel file of variable names and definitions has been specified
    '''
    lookup_df = None
    fname_lookup = settings['fname_lookup']

    if isfile(fname_lookup):

        # validate lookup Excel by checking required sheets
        # =================================================
        try:
            wb_obj = load_workbook(fname_lookup, data_only=True)
            sheet_names = wb_obj.sheetnames
            wb_obj.close()
            if APPNDX_A_SHEET not in sheet_names:
                fname_lookup = None

        except (PermissionError, BadZipFile) as err:
            print('Error: ' + str(err))
            fname_lookup = None

        if fname_lookup is not None:
            try:
                lookup_df = read_excel(fname_lookup, 'Appendix A')
                print('Successfully read lookup Excel file: ' + fname_lookup)
            except ValueError as err:
                print(ERROR_STR + str(err) + ' reading Excel file: ' + fname_lookup)
                fname_lookup = None
    else:
        print(ERROR_STR + ' lookup Excel file: ' + fname_lookup + '\n\tnot found - cannot run program ')
        fname_lookup = None

    settings['fname_lookup'] = fname_lookup
    if batch_flag:
        return lookup_df
    else:
        settings['lookup_df'] = lookup_df
        return 0