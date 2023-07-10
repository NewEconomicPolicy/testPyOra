#-------------------------------------------------------------------------------
# Name:        ora_excel_write.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Definitions:
#   spin_up
#
# Description:
#
#
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'ora_excel_write.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
import sys
from os.path import isfile, split, join
from os import remove
from glob import glob
from pandas import DataFrame, ExcelWriter, Series
from PyQt5.QtWidgets import QApplication
from warnings import filterwarnings

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment

from ora_classes_excel_write import A1SomChange, A2MineralN, A3SoilWater, A2aSoilNsupply, A2bCropNuptake, \
                A2cLeachedNloss, A2dDenitrifiedNloss, A2eVolatilisedNloss, A2fNitrification, \
                B1CropProduction, B1cNlimitation

from ora_lookup_df_fns import fetch_detail_from_varname

PREFERRED_LINE_WIDTH = 25000       # 100020 taken from chart_example.py     width in EMUs

from string import ascii_uppercase
ALPHABET = list(ascii_uppercase) + ['AA','AB','AC','AD']        # TODO: may need to extend
MAX_WDTH = 15                                   # width of column for table sheets
FIXED_WDTHS = {'A':13, 'B':6, 'C':7, 'D':11}    # period eg "steady state", year, month and crop name
WARN_STR = '*** Warning *** '

def generate_excel_outfiles(lggr, study, subarea, lookup_df, out_dir,  weather, complete_run, mngmnt_ss, mngmnt_fwd):
    '''

    '''
    study_full_name = study.study_name + ' ' + subarea      # typical study: Zerai Farm; subarea: Base line

    # concatenate weather into single entity
    # ======================================
    len_ss = len(weather.pettmp_ss['precip'])
    period_lst = len_ss*['steady state']

    len_fwd = len(weather.pettmp_fwd['precip'])
    period_lst += len_fwd*['forward run']

    precip_lst = weather.pettmp_ss['precip'] + weather.pettmp_fwd['precip']
    tair_lst = weather.pettmp_ss['tair'] + weather.pettmp_fwd['tair']
    pet_lst = weather.pettmp_ss['pet'] + weather.pettmp_fwd['pet']
    gdds_lst = weather.pettmp_ss['grow_dds'] + weather.pettmp_fwd['grow_dds']

    pettmp = {'period':period_lst, 'precip':precip_lst, 'tair':tair_lst, 'pet':pet_lst, 'grow_dds':gdds_lst}

    carbon_change, nitrogen_change, soil_water = complete_run
    nitrogen_change.additional_n_variables() # use existing data to populate additional fields
    print()     # cosmetic

    # make a safe name and
    # ===============
    fname = join(out_dir, study_full_name + '.xlsx')
    if isfile(fname):
        try:
            remove(fname)
        except PermissionError as err:
            print(err)
            return -1

    # use pandas to write to Excel
    # ============================
    writer = ExcelWriter(fname, engine='openpyxl')
    wb_map = {}  #  build a work book map for use when writing charts

    som_change_a1 = A1SomChange(pettmp, carbon_change, soil_water, mngmnt_ss, mngmnt_fwd)
    writer = _write_excel_out(lggr, 'A1 SOM change', som_change_a1, writer, wb_map)

    mineralN_a2 = A2MineralN(pettmp, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2 Mineral N', mineralN_a2, writer, wb_map)

    soilN_supply_a2a = A2aSoilNsupply(pettmp, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2a Soil N supply', soilN_supply_a2a, writer, wb_map)

    cropN_uptake_a2b = A2bCropNuptake(pettmp, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2b Crop N uptake', cropN_uptake_a2b, writer, wb_map)
     
    leachedN_loss_a2c = A2cLeachedNloss(pettmp, soil_water, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2c LeachedNloss', leachedN_loss_a2c, writer, wb_map)

    denit_Nloss_a2d = A2dDenitrifiedNloss(pettmp, carbon_change, nitrogen_change, soil_water)
    writer = _write_excel_out(lggr, 'A2d Denitrified N loss', denit_Nloss_a2d, writer, wb_map)

    volat_Nloss_a2e = A2eVolatilisedNloss(pettmp, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2e Volatilised N loss', volat_Nloss_a2e, writer, wb_map)

    nitrif_a2f = A2fNitrification(pettmp, nitrogen_change)
    writer = _write_excel_out(lggr, 'A2f Nitrification', nitrif_a2f, writer, wb_map)

    soil_water_a3 = A3SoilWater(pettmp, nitrogen_change, soil_water)
    writer = _write_excel_out(lggr, 'A3 Soil Water', soil_water_a3, writer, wb_map)

    crop_prodn_b1 = B1CropProduction(pettmp, soil_water, mngmnt_ss, mngmnt_fwd)
    writer = _write_excel_out(lggr, 'B1 Crop Production', crop_prodn_b1, writer, wb_map)

    n_limitation_b1c = B1cNlimitation(pettmp, carbon_change, nitrogen_change, soil_water, mngmnt_ss, mngmnt_fwd)
    writer = _write_excel_out(lggr, 'B1c Nitrogen Limitation', n_limitation_b1c, writer, wb_map)

    try:
        writer.close()
    except PermissionError as err:
        print(err)
        return -1

    # reopen Excel file and write charts
    # ==================================
    _generate_metric_charts(lggr, fname, lookup_df, wb_map)

    return

def _write_excel_out(lggr, sheet_name, out_obj, writer, wb_map):
    '''
    condition data before outputting
    '''
    func_name =  __prog__ +  ' write_excel_out'
    # filterwarnings('error')

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame()
    for var_name in out_obj.var_name_list:

        tmp_list = out_obj.sheet_data[var_name]
        if len(tmp_list) == 0:
            lggr.info(WARN_STR + 'no data for variable ' + var_name + ' in function ' + func_name)
            continue

        var_fmt = out_obj.var_formats[var_name]
        if var_fmt[-1] == 'f':
            ndecis = int(var_fmt[:-1])
            try:
                if var_name == 'crop_name':
                    data_frame[var_name] = Series(tmp_list)
                else:
                    data_frame[var_name] = Series([round(val, ndecis) for val in tmp_list])
                    '''
                    try:
                        data_frame[var_name] = Series([round(val, ndecis) for val in tmp_list])
                    except FutureWarning as warn:
                        print(warn)
                    '''
            except TypeError as err:
                print(err)
                return -1
        else:
            data_frame[var_name] = Series(tmp_list)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, index=False, freeze_panes=(1, 1))
    wb_map[sheet_name] = len(out_obj.var_name_list)

    return writer

def _generate_metric_charts(lggr, fname, lookup_df, wb_map):
    '''
    add charts to pre-existing Excel file
    '''
    func_name =  __prog__ + ' generate_charts'

    wb_obj = load_workbook(fname, data_only=True)
    ipos = 1
    for sheet_name in wb_obj.sheetnames:
        sheet = wb_obj[sheet_name]
        max_columns = wb_map[sheet_name]

        # reset column width and name
        # ===========================
        metric_dict = {}
        for ch in ALPHABET[3:max_columns]:
            cell_ref = ch + '1'
            metric = sheet[cell_ref].value
            if metric is None:
                mess = 'column: ' + ch + '\tsheet: ' + sheet_name + '\t' + 'has no metric at: ' + cell_ref
                lggr.info(WARN_STR + mess)
                continue

            metric_dict[cell_ref] = metric
            sheet.column_dimensions[ch].width = MAX_WDTH  # set the width of the column
            defn, units, out_format, pyora_disp = fetch_detail_from_varname(lookup_df, metric)
            if pyora_disp == metric:
                mess = 'column: ' + ch + '\tsheet: ' + sheet_name + '\t' + 'no lookup for metric: ' + metric
                lggr.info(WARN_STR + mess)

            sheet[cell_ref].value = pyora_disp
            sheet[cell_ref].alignment = Alignment(wrap_text = True, horizontal = 'center', vertical = 'center')

        for ch in FIXED_WDTHS:
            sheet.column_dimensions[ch].width = FIXED_WDTHS[ch]

        # adjust row heights
        # ==================
        for irow in range(sheet.max_row + 1):
            sheet.row_dimensions[irow].height = 18
        sheet.row_dimensions[1].height = 48

        # read and substitue field name
        # =============================

        # chart creation
        # ==============
        sheet_ref = sheet_name.split()[0]
        chart_sheet = wb_obj.create_sheet(sheet_ref + ' charts', ipos)
        ipos += 2
        nrow_chart = 10

        # generate charts for all metrics except for period, month and tstep
        # ==================================================================
        for col_indx in range(max_columns, 4, -1):               # ignore period, month and tstep fields
            metric_chart = LineChart()
            metric_chart.style = 13

            cell_ref = ALPHABET[col_indx - 1] + '1'
            try:
                metric = metric_dict[cell_ref]
            except KeyError as err:
                continue

            defn, units, out_format, pyora_disp = fetch_detail_from_varname(lookup_df, metric)
            metric_chart.title = defn
            metric_chart.y_axis.title = units
            metric_chart.x_axis.title = 'Time step'
            metric_chart.height = 10
            metric_chart.width = 20

            data = Reference(sheet, min_col = col_indx, min_row = 1, max_col = col_indx, max_row = sheet.max_row)
            metric_chart.add_data(data, titles_from_data = True)

            # Style the lines
            # ===============
            sref = metric_chart.series[0]
            sref.graphicalProperties.line.width = PREFERRED_LINE_WIDTH
            sref.graphicalProperties.line.solidFill = "FF0000"
            sref.smooth = True

            # now write to previously created sheet
            # =====================================
            chart_sheet.add_chart(metric_chart, "D" + str(nrow_chart))
            nrow_chart += 20

    try:
        wb_obj.active = 1   # which sheet to make active?
        wb_obj.save(fname)
        print('\tadded charts to: ' + fname)

    except PermissionError as err:
        print(str(err) + ' - could not save: ' + fname)

    QApplication.processEvents()

    return

def retrieve_output_xls_files(form, study_name = None):
    '''
    retrieve list of Excel files in the output directory
    '''
    if study_name is None:
        study_name = form.settings['study']

    out_dir = form.settings['out_dir']      # existence has been pre-checked in check_excel_input_fname
    xlsx_list = glob(out_dir + '/' + study_name + '*.xlsx')
    form.w_combo17.clear()
    if len(xlsx_list) > 0:
        form.w_disp_out.setEnabled(True)
        for out_xlsx in xlsx_list:
            dummy, short_fn = split(out_xlsx)
            form.w_combo17.addItem(short_fn)
    return
