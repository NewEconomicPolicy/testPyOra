'''
-------------------------------------------------------------------------------
 Name:        ora_excel_read_misc.py
 Purpose:     collection of miscellaneus functions required by the constructor
 Author:      Mike Martin
 Created:     23/02/2021
 Licence:     <your licence>
 Description:

-------------------------------------------------------------------------------
'''

__prog__ = 'ora_excel_read_misc.py'
__version__ = '0.0.1'
__author__ = 's03mm5'

from openpyxl import Workbook, load_workbook

fname = 'E:\\ORATOR\\study areas\\North Gondar (ETH)\\Pave\\vault\\FarmWthrMgmt_test.xlsx'
# order = {'Signature':0, 'Farm location':1, 'Weather':2, 'Subareas':3, 'A':4}

req_order = ['Signature', 'Farm location', 'Weather', 'Subareas', 'A']

def test_sheet_reordering():

    wb_obj = load_workbook(fname, data_only=True)
    sht_nmes = wb_obj.sheetnames
    my_order = [sht_nmes.index(sn) for sn in req_order]

    wb_obj._sheets = [wb_obj._sheets[indx] for indx in my_order]

    # wb_obj._sheets.sort(key = lambda ws: ws.title)
    # wb_obj.sheetnames
    # wb_obj.save('e:\\temp\\book_alphabetical.xlsx')

    wb_obj.save('e:\\temp\\FarmWthrMgmt_test.xlsx')

    wb_obj.close()

