#-------------------------------------------------------------------------------
# Name:        ora_utils_write_farm_sheets.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Description:
#
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'ora_utils_write_farm_sheets.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
from os.path import isfile, join, isdir
from os import makedirs, name as name_os

from pandas import DataFrame, ExcelWriter, Series
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from getpass import getuser
from datetime import datetime
from socket import gethostname

from ora_excel_read_misc import setup_sheet_data_dict, retrieve_hwsd_soil_recs, fetch_isda_soil_data
from getClimGenNC import ClimGenNC
from ora_wthr_misc_fns import associate_climate, fetch_csv_wthr
from ora_gui_misc_fns import simulation_yrs_validate, rotation_yrs_validate

from string import ascii_uppercase
ALPHABET = list(ascii_uppercase)

WARN_STR = '*** Warning *** '
GRANULARITY = 120       # notional resolution of 30 arc seconds i.e. 1/120 or ~0.0083333 degrees

FIXED_WDTHS = {}
FIXED_WDTHS['Signature'] = {'A':14, 'B':22}
FIXED_WDTHS['Farm location'] = {'A':28, 'B':15}
FIXED_WDTHS['Weather'] = {'A':13, 'B':9, 'C':9, 'D':9, 'E':9, 'F':9}    # period eg "steady state", year, month, precip and tair
FIXED_WDTHS['Subareas'] = {'A':10, 'B':25, 'C':10, 'D':10, 'E':10, 'F':8, 'G':8, 'H':8, 'I':8, 'J':8, 'K':8, 'L':8 }
FIXED_WDTHS['Livestock'] = {'A':15, 'B':25, 'C':25, 'D':25, 'E':25, 'F':25, 'G':25 }
NFEED_TYPES = 5

SHEET_NAMES = {'sign': 'Signature', 'lctn': 'Farm location', 'wthr':'Weather', 'sbas':'Subareas', 'lvstck':'Livestock'}
FLDS_ROW_HGHT = 20
STD_ROW_HGHT = 18
SALINITY = 0.1  # TODO

# clay, sand, silt and organic carbon are % by weight
# ===================================================
SOIL_METRIC_DESCRS = {'area': 'Area (ha)', 'depth': 'Soil depth (cm)',
        'clay': 'Clay content %', 'sand': 'Sand content %', 'silt':'Silt content %', 'oc':'Organic Carbon content %',
                       'bulk': 'Bulk Density g/cm**3', 'ph': 'PH', 'salinity': 'salinity (EC 1:5)'}

def _fetch_actual_year_lists(strt_yr_ss, end_yr_ss, nmnths_ss, strt_yr_fwd, end_yr_fwd):
    '''
    return list of years for both steady state and forward runs
    '''
    yr_lst = [] # steady state
    for year in range(strt_yr_ss, end_yr_ss):
        yr_lst += 12 * [year]

    actl_yr_lst = yr_lst[:nmnths_ss]

    yr_lst = [] # forward run
    for year in range(strt_yr_fwd, end_yr_fwd):
        yr_lst += 12 * [year]

    actl_yr_lst += yr_lst

    return actl_yr_lst

def _fetch_existing_subarea_sheets(fname_run):
    '''
    return list of subarea sheets
    '''
    wb_obj = load_workbook(fname_run, data_only=True)
    sheet_names = wb_obj.sheetnames
    wb_obj.close()

    sba_sheets = []
    for sht_nm in sheet_names:
        if sht_nm in ALPHABET:
            sba_sheets.append(sht_nm)

    return sba_sheets

def _write_xls_subarea_summary_sheet(sheet_name, exstng_sbas, form, soil_recs, writer):
    '''
    write sheet recording state of subarea widgets

    TODO: for future
    subarea_field_map = {'sba': 'Subarea', 'descr': 'Description', 'irrig': 'Irrigation (mm)', 'rota': 'Rotation (yrs)',
                         'area': 'Area (ha)'}
    Irrigation (mm)': [],  'Rotation (yrs)':
    '''
    # condition data
    # ==============
    t_depth, t_bulk, t_ph, t_clay, t_silt, t_sand, t_c_prcnt = soil_recs[0]

    subarea_dict = {'Subarea': [], 'Description': [], 'Irrig (mm)': [],  'Rota (yrs)': [], 'Area (ha)': [],
                    't_clay':[], 't_sand':[], 't_silt':[],  't_oc':[], 't_bulk':[], 't_ph':[], 'salin':[] }

    for sba_indx in ALPHABET[:form.settings['nsubareas']]:
        subarea_dict['Subarea'].append(sba_indx)
        descr = form.w_sba_descrs[sba_indx].text()
        if descr == '':
            # supply description when sheet for subarea exists but description is blank
            # =========================================================================
            if sba_indx in exstng_sbas:
                descr = sba_indx
        else:
            # sheet for subarea must exist otherwise nullify description
            # ==========================================================
            if sba_indx not in exstng_sbas:
                descr = ''

        subarea_dict['Description'].append(descr)
        # TODO: subarea_dict['Irrigation (mm)'].append(int(form.w_typ_irri[sba_indx].text()))
        subarea_dict['Irrig (mm)'].append(50)
        nyrs_rota = rotation_yrs_validate(form.w_nrota_ss[sba_indx])
        subarea_dict['Rota (yrs)'].append(nyrs_rota)
        subarea_dict['Area (ha)'].append(float(form.w_areas[sba_indx].text()))
        subarea_dict['t_clay'].append(t_clay)
        subarea_dict['t_sand'].append(t_sand)
        subarea_dict['t_silt'].append(t_silt)
        subarea_dict['t_oc'].append(t_c_prcnt)
        subarea_dict['t_bulk'].append(t_bulk)
        subarea_dict['t_ph'].append(t_ph)
        subarea_dict['salin'].append(SALINITY)

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame(subarea_dict)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, index=False, freeze_panes=(1, 1))

    return writer

def make_or_update_farm(form):
    '''
    use pandas to write to Excel
    if new_runfile_flag is set to True then the a new run file will be created or, if False, edited
    return code is -1 for fail and 0 for success together with new_runfile_flag
    '''

    # weather
    # =======
    snglPntFlag = True
    num_band = 0
    hwsd = SetResolution()
    lggr = form.lggr

    wthr_rsrc = form.w_combo30w.currentText()
    fut_clim_scen = form.w_combo30.currentText()
    strt_yr_ss = int(form.w_combo29s.currentText())
    strt_yr_fwd = int(form.w_combo31s.currentText())
    study_dir = join(form.settings['study_area_dir'], form.w_combo00.currentText())
    farm_dir = join(study_dir, form.w_farm_name.text())
    if not isdir(farm_dir):
        makedirs(farm_dir)

    lat = float(form.w_lat.text())
    lon = float(form.w_lon.text())
    bbox_aoi = list([lon - 0.01, lat - 0.01, lon + 0.01, lat + 0.01])

    # soil - if lat/lon are out of area for iSDA then use HWSD
    # ========================================================
    use_isda_flag = form.w_use_isda.isChecked()
    if use_isda_flag:
        use_isda_flag, soil_recs = fetch_isda_soil_data(lggr, lat, lon)

    if not use_isda_flag:
        soil_recs = retrieve_hwsd_soil_recs(lggr, form.settings['hwsd_dir'], lat, lon)

    gran_lat = int(round((90.0 - lat) * hwsd.granularity))
    gran_lon = int(round((180.0 + lon) * hwsd.granularity))
    site_rec = list([gran_lat, gran_lon, lat, lon, None, None])

    # weather
    # =======
    nyrs_ss, nyrs_fwd = simulation_yrs_validate(form.w_nyrs_ss, form.w_nyrs_fwd)
    end_yr_ss = strt_yr_ss + nyrs_ss
    end_yr_fwd = strt_yr_fwd + nyrs_fwd

    if form.wthr_sets is None or form.w_use_csv.isChecked():
        wthr_src = 'CSV file'
        csv_fn = form.w_csv_fn.text()
        nmnths_ss, pettmp_ss, nmnths_fwd, pettmp_fwd = fetch_csv_wthr(csv_fn, nyrs_ss, nyrs_fwd)
    else:
        wthr_src = 'NC dataset'
        climgen = ClimGenNC(lggr, form.wthr_sets, wthr_rsrc, fut_clim_scen,
                                                                        strt_yr_ss, end_yr_ss, strt_yr_fwd, end_yr_fwd)
        # generate weather dataset indices which enclose the AOI
        # ======================================================
        aoi_indices_fut, aoi_indices_hist = climgen.genLocalGrid(bbox_aoi, hwsd, snglPntFlag, num_band)

        pettmp_fut = climgen.fetch_cru_future_NC_data(aoi_indices_fut, num_band)
        pettmp_hist = climgen.fetch_cru_historic_NC_data(aoi_indices_hist, num_band)

        nmnths_ss, pettmp_ss, nmnths_fwd, pettmp_fwd = associate_climate(site_rec, climgen, pettmp_hist, pettmp_fut)

    print('Retrieved {} months of weather data for simulation run from {}'.format(nmnths_ss + nmnths_fwd, wthr_src))
    '''
    this section is taken from PyOrator ..\InitInptsRslts\ora_excel_write.py
    '''
    # concatenate weather into single entity
    # ======================================
    len_ss = len(pettmp_ss['precip'])
    period_lst = len_ss * ['steady state']

    len_fwd = len(pettmp_fwd['precip'])
    period_lst += len_fwd * ['forward run']

    precip_lst = pettmp_ss['precip'] + pettmp_fwd['precip']
    tair_lst = pettmp_ss['tair'] + pettmp_fwd['tair']

    actl_yr_lst = _fetch_actual_year_lists(strt_yr_ss, end_yr_ss, nmnths_ss, strt_yr_fwd, end_yr_fwd)

    pettmp = {'period': period_lst, 'precip': precip_lst, 'tair': tair_lst, 'actl_yr': actl_yr_lst}

    # make a safe name
    # ================
    exstng_sbas = []
    fname_run = join(farm_dir, form.settings['fname_run'])
    if isfile(fname_run):
        new_runfile_flag =  False
        exstng_sbas = _fetch_existing_subarea_sheets(fname_run)
        _delete_lctn_wthr_sheets(fname_run)     # delete all sheets except Signature
        writer = ExcelWriter(fname_run, mode='a', if_sheet_exists='replace')
    else:
        new_runfile_flag = True
        writer = ExcelWriter(fname_run)

    wthr = WeatherSheet(pettmp)
    if new_runfile_flag:
        writer = _write_excel_signature(SHEET_NAMES['sign'], writer)
    writer = _write_excel_location(SHEET_NAMES['lctn'], form, writer)
    writer = _write_excel_weather(SHEET_NAMES['wthr'], wthr, writer)
    writer = _write_excel_livestock(SHEET_NAMES['lvstck'], form, writer)
    area_ha = None

    writer = _write_xls_subarea_summary_sheet(SHEET_NAMES['sbas'], exstng_sbas, form, soil_recs, writer)
    try:
        writer.save()
        writer.close()
    except PermissionError as err:
        print(err)
        return -1, None

    # make sheets more readable
    # =========================
    _adjust_excel_workbook(fname_run)

    return 0, new_runfile_flag

def _adjust_excel_workbook(fname_run):
    '''
    improve readability of sheets and record modification time in Signature sheet
    '''
    greenFill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    wb_obj = load_workbook(fname_run, data_only=True)
    for sheet_name in wb_obj.sheetnames:
        sheet = wb_obj[sheet_name]

        if sheet_name in ALPHABET:   # subarea sheets are adjusted in management functions
            continue

        if sheet_name not in SHEET_NAMES.values():
            print(WARN_STR + 'unrecognised sheet ' + sheet_name + ' in workbook ' + fname_run)
            continue

        fixed_wdths = FIXED_WDTHS[sheet_name]

        # adjust fill color and alignment for headers row only
        # ====================================================
        for col_indx in fixed_wdths:
            cell_ref = col_indx + '1'
            sheet[cell_ref].fill = greenFill
            sheet[cell_ref].alignment = Alignment(horizontal='center')

        # adjust alignment for first column only
        # ======================================
        if sheet_name == 'Subareas':
            algnmnt = 'center'
        else:
            algnmnt = 'left'

        for irow in range(2, sheet.max_row + 1):
            cell_ref = 'A' + str(irow)
            sheet[cell_ref].alignment = Alignment(horizontal = algnmnt)

        # adjust row heights and column widths
        # ====================================
        for irow in range(sheet.max_row + 1):
            sheet.row_dimensions[irow].height = STD_ROW_HGHT

        if sheet_name == 'Signature':
            modify_date = datetime.now().strftime("%b %d %Y %H:%M:%S")
            sheet['B6'].value = getuser()
            sheet['B7'].value = modify_date

        elif sheet_name == 'Weather':

            # adjust notional and actual year
            # ===============================
            for clmn in list(['B', 'F']):
                for irow in range(2, sheet.max_row + 1):
                    cell_ref = clmn + str(irow)
                    sheet[cell_ref].alignment = Alignment(horizontal='center')

        # adjust column widths
        # ====================
        for ch in fixed_wdths:
            sheet.column_dimensions[ch].width = fixed_wdths[ch]

    try:
        wb_obj.active = 0   # which sheet to make active?
        wb_obj.save(fname_run)
        print('\tadded sheets to: ' + fname_run)

    except PermissionError as err:
        print(str(err) + ' - could not save: ' + fname_run)

    return

def _write_excel_weather(sheet_name, out_obj, writer):
    '''
    condition data before outputting
    '''
    func_name =  __prog__ +  ' write_excel_out'

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame()
    for var_name in out_obj.var_name_list:

        tmp_list = out_obj.sheet_data[var_name]

        var_fmt = out_obj.var_formats[var_name]
        if var_fmt[-1] == 'f':
            ndecis = int(var_fmt[:-1])
            try:
                if var_name == 'crop_name':
                    data_frame[var_name] = Series(tmp_list)
                else:
                    data_frame[var_name] = Series([round(val, ndecis) for val in tmp_list])
            except TypeError as err:
                print(err)
                return -1
        else:
            data_frame[var_name] = Series(tmp_list)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, index=False, freeze_panes=(1, 1))
    return writer

def _write_excel_signature(sheet_name, writer):
    '''
    this sheet edited when changing run file
    '''
    username = getuser()
    create_date = datetime.now().strftime("%b %d %Y %H:%M:%S")
    hostname = gethostname()

    signature = {'Attribute': ['Workstation', 'OS', 'Created by', 'Create date', 'Modified by', 'Modified date'],
                     'Value': [hostname, name_os, username, create_date, '', '']}

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame(signature)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, index=False, freeze_panes=(1, 1))

    return writer

def _write_excel_location(sheet_name, form, writer):
    '''
    TODO: validate data before outputting
    '''
    latitude = float(form.w_lat.text())
    longitude = float(form.w_lon.text())
    subdist = form.w_subdist.text()
    farm_name = form.w_farm_name.text()
    farm_desc = form.w_farm_desc.text()
    farm_area = float(form.w_area.text())
    prcnt_subdist = float(form.w_prcnt.text())

    farm_location = {'Attribute': ['Subdistrict name', 'Farm name', 'Latitude', 'Longitude',
                                        'Area (ha)','% of subdistrict','Description'],
                     'Value': [subdist, farm_name, latitude, longitude, farm_area, prcnt_subdist, farm_desc]}

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame(farm_location)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, index=False, freeze_panes=(1, 1))

    return writer

def _write_excel_livestock(sheet_name, form, writer):
    '''
    condition data before outputting - uses widgets: w_bought_in, w_feed_qties, w_feed_types, w_numbers, w_strtgs
    '''
    func_name =  __prog__ +  ' _write_excel_livestock'

    anml_typs = form.anml_prodn.gnrc_anml_types

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame()
    data_frame['dscrs'] = Series(form.lvstck_row_dscrs)

    for anml in anml_typs:
        tmp_list = [anml_typs[anml]]    # field description
        try:
            num_anmls = float(form.w_numbers[anml].text())
        except ValueError as err:
            num_anmls = 0
        tmp_list.append(num_anmls)
        tmp_list.append(form.w_strtgs[anml].currentText())

        for findx in range(NFEED_TYPES):
            fd_typ = str(findx + 1)
            tmp_list.append(form.w_feed_types[anml][fd_typ].currentText())
            try:
                feed_qty = float(form.w_feed_qties[anml][fd_typ].text())
            except ValueError as err:
                feed_qty = 0
            tmp_list.append(feed_qty)

        tmp_list.append(form.w_bought_in[anml].text())

        data_frame[anml] = Series(tmp_list)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, sheet_name, header=False, index=False, freeze_panes=(1, 1))

    return writer

def _delete_lctn_wthr_sheets(fname_run, sheet_del = 'all'):
    '''
    sheets_del: sheets to delete, all or a key in  SHEET_NAMES
    Signature sheet is not deleted
    '''
    func_name =  __prog__ + ' _remove_excel_sheets'

    wb_obj = load_workbook(fname_run, data_only=True)
    for sheet_name in SHEET_NAMES.values():
        if sheet_name == 'Signature':
            continue
        else:
            if sheet_del == 'all' or sheet_name == sheet_del:
                try:
                    del wb_obj[sheet_name]
                except KeyError as err:
                    print('Could not delete sheet ' + sheet_name)
    try:
        wb_obj.save(fname_run)
    except PermissionError as err:
        print(str(err) + ' - could not save: ' + fname_run)

    wb_obj.close()

    return

class WeatherSheet(object, ):

    def __init__(self, pettmp):
        '''
        B1      TODO: removed 'days_month': 'd' from var_format_dict - don't know why it was there
        '''
        self.title = 'Weather'

        var_format_dict = {'period': 's',  'year':'d', 'month': 'd', 'precip': '2f', 'tair': '2f', 'actl_yr': 'd'}

        sheet_data, var_name_list, exclusion_list = setup_sheet_data_dict(pettmp, var_format_dict)

        sheet_data['precip'] = [float('{:=.2f}'.format(val)) for val in pettmp['precip']]       # TODO: inefficient
        sheet_data['tair'] = [float('{:=.2f}'.format(val)) for val in pettmp['tair']]
        sheet_data['actl_yr'] = pettmp['actl_yr']

        self.sheet_data = sheet_data
        self.var_name_list = var_name_list
        self.var_formats = var_format_dict

class SetResolution(object):
    """
    a minimal class to ensure compatiblity with Global Ecosse modules
    sets resolution
    """
    def __init__(self):

        self.granularity = GRANULARITY