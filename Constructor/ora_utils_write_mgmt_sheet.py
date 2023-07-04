#-------------------------------------------------------------------------------
# Name:        ora_utils_write_mgmt_sheets.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Description:
#
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'ora_utils_write_mgmt_sheets.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
from pandas import DataFrame, ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, PatternFill, Font, Border
from getpass import getuser
from datetime import datetime

from ora_excel_read_misc import setup_sheet_data_dict

ERROR_STR = '*** Error *** '
WARN_STR = '*** Warning *** '
SHEET_NAMES = {'sign': 'Signature', 'lctn': 'Farm location', 'wthr':'Weather'}

# TODO: combine fixed widths with var_format_dict
# period, year, month,	Crop Yield, Fert type, Fert N, OW type, OW amount, irrigation
# ===================================================================================
FIXED_WDTHS_SBA = {'A':13, 'B':7, 'C':7, 'D':20, 'E':8, 'F':10, 'G':8, 'H':15, 'I':14, 'J':11}
FLDS_ROW_HGHT = 20
STD_ROW_HGHT = 18

def write_mgmt_sht(fname, subarea, sba_descr, nmnths_ss, nmnths_fwd, rota_dict, overwrite_flag = True):
    '''
    use pandas to create a data frame and write a management sheet to Excel
    if overwrite_flag is set to True then the existing sheet will deleted then be rewritten
    '''
    period_lst = nmnths_ss * ['steady state']
    period_lst += nmnths_fwd * ['forward run']

    mngmnt = _rework_rota_to_mngmnt(period_lst, rota_dict)

    _remove_subarea_sheet(fname, subarea)
    writer = ExcelWriter(fname, mode='a')

    # create data frame from dictionary
    # =================================
    data_frame = DataFrame(mngmnt)

    # output frame and record number of columns
    # =========================================
    data_frame.to_excel(writer, subarea, index=False, freeze_panes=(1, 1))

    try:
        writer.close()
    except PermissionError as err:
        print(err)
        return -1

    # reopen, adjust sheets and refresh subareas table
    # ================================================
    _adjust_subarea_sheets(fname, subarea)
    print('\tadded sheet ' + subarea + ' and adjusted sheets in: ' + fname)

    return

def _blank_zeroes(sheet, val_col, irow):
    '''
       blank zeros
    '''
    cell_val = val_col + str(irow)
    if sheet[cell_val].value == 0:
        sheet[cell_val].value = ''

    return

def _blank_empty_attrib_value(sheet, attr_col, amnt_col, irow):
    '''
    blank empty attribute and associated value e.g. organic waste and amount
    '''
    cell_type = attr_col + str(irow)
    cell_val = amnt_col + str(irow)
    if sheet[cell_type].value == 'None':
        sheet[cell_type].value = ''
        sheet[cell_val].value = ''

    return

def _adjust_subarea_sheets(fname, subarea):
    '''
    update signature sheet and improve readability of this subarea sheet by adjusting column widths and alignments
    possible future modification to make more readable:
        remove zeros from yield, Fert N, OW amount and irrigation, columns E, G, I and J
        remove No crop from Crop, column D
        remove Nones from Fert type and OW type, columns F and H
    '''
    # greenFill = PatternFill(start_color= 'lightGreen', end_color='lightGreen', fill_type='solid')
    # greenFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

    greenFill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    siggy = SHEET_NAMES['sign']
    fixed_wdths = FIXED_WDTHS_SBA

    wb_obj = load_workbook(fname, data_only=True)

    # unlikely to happen
    # ==================
    if siggy not in wb_obj.sheetnames:
        print(ERROR_STR + 'sheet ' + siggy + ' not in workbook ' + fname)
        wb_obj.close()
        return

    # record modification time in Signature sheet
    # ===========================================
    sheet = wb_obj[siggy]
    modify_date = datetime.now().strftime("%b %d %Y %H:%M:%S")
    sheet['B6'].value = getuser()
    sheet['B7'].value = modify_date

    # modify this subarea sheet
    # =========================
    sheet = wb_obj[subarea]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # adjust row heights
    # ==================
    sheet.row_dimensions[1].height = FLDS_ROW_HGHT     # field names
    for irow in range(sheet.max_row + 1):
        sheet.row_dimensions[irow].height = STD_ROW_HGHT

    for ch in fixed_wdths:
        sheet.column_dimensions[ch].width = fixed_wdths[ch]

    # adjust values
    # =============
    for irow in range(2, max_row + 1):
        cell_crop = 'D' + str(irow)
        if sheet[cell_crop].value == 'No crop':
            sheet[cell_crop].value = ''
        else:
            sheet[cell_crop].fill = greenFill

        # fert and ow
        # ===========
        _blank_zeroes(sheet, 'E', irow)                     # yield
        _blank_empty_attrib_value(sheet, 'F', 'G', irow)    # fertiliser
        _blank_empty_attrib_value(sheet, 'H', 'I', irow)    # OW
        _blank_zeroes(sheet, 'J', irow)                     # irrigation

    try:
        wb_obj.active = 0  # which sheet to make active?
        wb_obj.save(fname)
    except PermissionError as err:
        print(str(err) + ' - could not save: ' + fname)

    wb_obj.close()

    return

def _cnvrt_list_to_numbers(val_lst):
    '''

    '''
    rslt_lst = []
    for val in val_lst:
        try:
            rslt = float(val)
        except ValueError:
            rslt = 0

        rslt_lst.append(rslt)

    return rslt_lst

def _rework_rota_to_mngmnt(period_lst, rota_dict):
    '''
    expand rotation to management for steady state and forward run
    '''
    rota_len = len(rota_dict['crop_names'])
    nperiods = int(len(period_lst)/rota_len)

    yld_typcls = _cnvrt_list_to_numbers(rota_dict['yld_typcls'])
    fert_n_amnts =  _cnvrt_list_to_numbers(rota_dict['fert_n_amnts'])
    ow_amnts = _cnvrt_list_to_numbers(rota_dict['ow_amnts'])
    irrigs = _cnvrt_list_to_numbers(rota_dict['irrigs'])

    crop_lst = []; yld_lst = []; fert_typ_lst = []; fert_n_lst = []; ow_typ_lst = []; ow_amnt_lst = []; irrig_lst = []

    for ic in range(nperiods):
        crop_lst += rota_dict['crop_names']
        yld_lst += yld_typcls
        fert_typ_lst += rota_dict['fert_typs']
        fert_n_lst += fert_n_amnts
        ow_typ_lst += rota_dict['ow_typs']
        ow_amnt_lst += ow_amnts
        irrig_lst += irrigs

    all_data = (crop_lst, yld_lst, fert_typ_lst, fert_n_lst, ow_typ_lst, ow_amnt_lst, irrig_lst)

    var_format_dict = {'period': 's', 'year': 'd', 'month': 'd', 'Crop': 's', 'Yield': '3f',
                            'Fert type': 's', 'Fert N': '2f', 'OW type': 's', 'OW amount': '2f', 'irrigation': '2f'}

    period_dict = {'period': period_lst}
    mngmnt, var_name_lst, exclusion_list = setup_sheet_data_dict(period_dict, var_format_dict)

    for metric, metric_lst in zip(var_name_lst[3:], all_data):
        mngmnt[metric] = metric_lst

    return mngmnt

def _remove_subarea_sheet(fname, subarea):
    '''
    remove subarea sheet from pre-existing Excel file
    '''
    wb_obj = load_workbook(fname, data_only=True)
    if subarea in wb_obj.sheetnames:
        del wb_obj[subarea]
        try:
            wb_obj.save(fname)
        except PermissionError as err:
            print(str(err) + ' - could not save: ' + fname)

    wb_obj.close()

    return
