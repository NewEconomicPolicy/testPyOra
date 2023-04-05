'''
-------------------------------------------------------------------------------
 Name:        ora_excel_read_misc.py
 Purpose:     collection of miscellaneus functions required by the constructor
 Author:      Mike Martin
 Created:     23/02/2021
 Licence:     <your licence>
 Description:

-------------------------------------------------------------------------------
'''

__prog__ = 'ora_excel_read_misc.py'
__version__ = '0.0.1'
__author__ = 's03mm5'

from os.path import join, split, dirname
from glob import glob
from openpyxl import load_workbook
from pandas import DataFrame
import requests

import hwsd_bil

ERROR_STR = '*** Error *** '
SHEET_NAMES = {'sign': 'Signature', 'lctn': 'Farm location', 'wthr':'Weather', 'sbas':'Subareas',
                                                                                                'lvstck':'Livestock'}
ANML_ABBREVS = ['catdry','catmt','rumdry','rummt','pigs','pltry']
MONTH_NAMES_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

ISDA_URL = 'https://api.isda-africa.com/v1/soilproperty'
SOIL_DATA_API_KEY = 'AIzaSyDdmmXVqnAVXRE28meqX2XMQxzRox_25RM'

def fetch_isda_soil_data(lggr, lat, lon):
    '''
    original request string from Dave McBey:
          f'https://api.isda-africa.com/v1/soilproperty?key={soil_data_api_key}&lat={lat}&lon={lon}')

    It is possible to request specific properties: '&lat={}&lon={}&property=ph&depth=0-20'.format(lat, lon)
    '''
    soil_recs = []
    request_str = ISDA_URL + '?key=' + SOIL_DATA_API_KEY + '&lat={}&lon={}'.format(lat, lon)
    soil_data = requests.get(request_str)
    if soil_data.ok:

        # create a soil file and a formatted string of the JSON object
        # ============================================================
        """soil_file = 'E:\\temp\\soil.json'
        if isfile(soil_file):
            remove(soil_file)

        with open(soil_file, 'w') as fsetup:
            json.dump(soil_data.json(), fsetup, indent=2, sort_keys=True)"""

        t_depth = 20
        isda_data = soil_data.json()
        isda = isda_data['property']
        t_c_prcnt = round(isda['carbon_organic'][0]['value']['value']/10.0, 4)
        t_bulk = isda['bulk_density'][0]['value']['value']
        t_ph = isda['ph'][0]['value']['value']
        t_clay = isda['clay_content'][0]['value']['value']
        t_silt = isda['silt_content'][0]['value']['value']
        t_sand  = isda['sand_content'][0]['value']['value']

        soil_recs.append([t_depth, t_bulk, t_ph, t_clay, t_silt, t_sand, t_c_prcnt])

    return soil_data.ok, soil_recs

def retrieve_hwsd_soil_recs(lggr, hwsd_dir, lat, lon):
    '''
    extract required metrics from the HWSD database
    '''
    try:
        hwsd = hwsd_bil.HWSD_bil(lggr, hwsd_dir)
    except:
        return

    nvals_read = hwsd.read_bbox_mu_globals([lon, lat], snglPntFlag = True)
    lggr.info('Read {} value from HWSD'.format(nvals_read))

    # retrieve dictionary mu_globals and number of occurrences
    # ========================================================
    mu_globals = hwsd.get_mu_globals_dict()
    if mu_globals is None:
        print('No soil records for this area\n')
        return

    # create and instantiate a new class NB this stanza enables single site
    # ==================================
    hwsd_soil_recs = hwsd.get_soil_recs(sorted(mu_globals.keys()))
    ''' 
    an HWSD soil has a mandatory top and optional sub layer          
        t_c = t_bulk * (t_oc/100.0) * 30 * (100000000/1000.0)   # t_oc is percentage carbon
        64740  # C content [kgC/ha]
        1.50   # Bulk density [g/cm3]
        8.3    # pH
        50.0   # % clay by weight
        21.0   # % silt by weight
        29.0   # % sand by weight for this
    '''
    t_depth = 30
    for key in hwsd_soil_recs:
        t_c, t_bulk, t_ph, t_clay, t_silt, t_sand = hwsd_soil_recs[key][0][:6]
    t_c_prcnt = round(t_c / (t_bulk * t_depth * 1000), 6)

    mess = 'Retrieved {} values  of HWSD grid consisting of {} rows and {} columns: ' \
           '\n\tnumber of unique mu_globals: {}'.format(nvals_read, hwsd.nlats, hwsd.nlons, len(mu_globals))
    lggr.info(mess)
    print(mess)

    soil_recs = []
    soil_recs.append([t_depth, t_bulk, t_ph, t_clay, t_silt, t_sand, t_c_prcnt])

    return soil_recs

def read_farm_wthr_sbsa_xls_file(form, run_xls_fn):
    '''
    check required sheets are present
    '''

    # set widgets belonging to subareas to defaults
    # =============================================
    for sba in form.w_nrota_ss:
        form.w_nrota_ss[sba].setText(str(form.settings['nrota_yrs_dflt']))
        form.w_areas[sba].setText(str(form.settings['areas_dflt']))
        form.w_sba_descrs[sba].setText('')
        form.w_ss_mgmt[sba].setEnabled(False)

    ret_var = None
    wb_obj = load_workbook(run_xls_fn, data_only=True)

    rqrd_sheet = SHEET_NAMES['sbas']
    if rqrd_sheet in wb_obj.sheetnames:
        sbas_sht = wb_obj[rqrd_sheet]
        df = DataFrame(sbas_sht.values, columns=['Subarea','Description','Irrig','Rotation','Areas',
                                                 't_clay', 't_sand', 't_silt',  't_oc', 't_bulk', 't_ph', 'salin'])

        # check subarea sheet exists for a non-blank description
        # ======================================================
        for row in df.values[1:]:
            sba, descr, irrig, nrota_yrs, area = row[:5]

            # trap for missing description
            # ============================
            if sba in wb_obj.sheetnames and descr is None:
                descr = sba + 'dummy'
                WARN_STR = '*** Warning *** '
                mess = WARN_STR + 'management sheet for subarea ' + sba + ' is present'
                mess += ' but description is missing from ' + rqrd_sheet + ' sheet - will use ' + descr
                print(mess)

            if descr is not None:
                if sba in wb_obj.sheetnames:
                    form.w_sba_descrs[sba].setText(descr)
                    # form.w_typ_irri[sba].setText(str(irrig))
                    form.w_nrota_ss[sba].setText(str(nrota_yrs))
                    form.w_areas[sba].setText(str(area))
                    form.w_ss_mgmt[sba].setEnabled(True)
                else:
                    mess = WARN_STR
                    mess += 'discarded description ' + descr + ' for subarea ' + sba + ' - no associated sheet present'
                    print(mess)
    else:
        print('Sheet ' + rqrd_sheet + ' not present in ' + run_xls_fn)

    # populate Livestock page
    # =======================
    anml_abbrevs = ANML_ABBREVS
    NFEED_TYPES = 5
    rqrd_sheet = SHEET_NAMES['lvstck']
    if rqrd_sheet in wb_obj.sheetnames:
        lvstck_sht = wb_obj[rqrd_sheet]
        df = DataFrame(lvstck_sht.values, columns = ['descr'] + anml_abbrevs) # ['descr'] +
        irow = 1
        for anml, nmbr in zip(anml_abbrevs, df.values[irow][1:]):      # numbers of animals
            if nmbr is not None:
                form.w_numbers[anml].setText(str(nmbr))
        irow += 1
        for anml, strtg in zip(anml_abbrevs, df.values[irow][1:]):     # strategies
            form.w_strtgs[anml].setCurrentText(strtg)

        for findx in range(NFEED_TYPES):
            irow += 1
            fd_typ = str(findx + 1)
            for anml, feed_typ in zip(anml_abbrevs, df.values[irow][1:]):  # feed type
                form.w_feed_types[anml][fd_typ].setCurrentText(feed_typ)

            irow += 1
            for anml, feed_qty in zip(anml_abbrevs, df.values[irow][1:]):  # feed quantity as a percentage
                form.w_feed_qties[anml][fd_typ].setText(str(feed_qty))

        irow += 1
        for anml, pcnt in zip(anml_abbrevs, df.values[irow][1:]):     # percentage bought in
            if pcnt is not None:
                form.w_bought_in[anml].setText(str(pcnt))
    else:
        print('Sheet ' + rqrd_sheet + ' not present in ' + run_xls_fn)

    wb_obj.close()

    return ret_var

def clear_farm_fields(form):
    '''

    '''
    form.w_farm_desc.setText('')
    form.w_prcnt.setText('')
    form.w_subdist.setText('')
    form.w_area.setText(str(''))
    form.w_lat.setText(str(''))
    form.w_lon.setText(str(''))

def validate_farm_var_fields(form):
    '''
    invoked when Save farm push button is changed
    '''
    mess = ERROR_STR
    validate_flag = True

    farm_name = form.w_farm_name.text()
    if len(farm_name) == 0:
        validate_flag = False
        mess += 'Farm name must not be empty\t'

    try:
        lat = float(form.w_lat.text())
    except ValueError:
        validate_flag = False
        mess += 'Latitude must be a float\t'

    try:
        lon = float(form.w_lon.text())
    except ValueError:
        validate_flag = False
        mess += 'Longitude must be a float\t'

    try:
        farm_area = float(form.w_area.text())
    except ValueError:
        validate_flag = False
        mess += 'Farm area must be a float\t'

    try:
        prcnt_subdist = float(form.w_prcnt.text())
    except ValueError:
        validate_flag = False
        mess += 'Percentage of sub-district must be a float\t'

    if not validate_flag:
        print(mess)

    return validate_flag

def get_mnth_yr_names(nyears):
    '''
    returns unique month year strings
    '''
    mnth_keys = []
    for yr in range(nyears):
        str_yr = str(yr + 1)
        for mnth in MONTH_NAMES_SHORT:
            mnth_keys.append(mnth + ' ' + str_yr)

    return mnth_keys

def identify_study_areas(form, study_area_dir, fname_run):

    study_dirs = glob(study_area_dir + '/*(*)')  # TODO improve

    study_areas_valid = []
    if len(study_dirs) > 0:

        for study_dir in study_dirs:
            nfarms = identify_farms_for_study(form, study_dir, fname_run)
            if len(nfarms) >= 0:
                study_areas_valid.append(study_dir)

    return study_areas_valid

def identify_farms_for_study(form, study_dir = None, fname_run = None):
    '''
    is called at start up and when user creates a new farm project
    farms = {'New farm':''}
    '''
    if form is not None:
        if hasattr(form, 'w_combo00'):
            study = form.w_combo00.currentText()
        else:
            study = form.w_tab_wdgt.w_combo00.currentText()

        study_dir = join(form.settings['study_area_dir'], study)
        fname_run = form.settings['fname_run']

    farm_files = glob(study_dir + '\\*\\' + fname_run)
    farms = {}
    for fname in farm_files:
        farm_name = split(dirname(fname))[1]
        if farm_name != '' :
            farms[farm_name] = fname

    return farms

def check_sheets_for_farms(form):
    '''
    is called at start up and when user creates a new farm project
    farms = {'New farm':''}
    '''

    study = form.w_combo00.currentText()
    study_dir = join(form.settings['study_area_dir'], study)
    fname_run = form.settings['fname_run']

    farm_files = glob(study_dir + '\\*\\' + fname_run)
    farms = {}
    for fname in farm_files:
        farm_name = split(dirname(fname))[1]
        if farm_name != '':

            # check mandatory sheets are present
            # ==================================
            farm_flag = True
            wb_obj = load_workbook(fname, data_only=True)
            for sheet_name in SHEET_NAMES.values():
                if sheet_name not in wb_obj.sheetnames:
                    print(ERROR_STR + 'sheet ' + sheet_name + ' not present in ' + farm_name)
                    farm_flag = False

            if farm_flag:
                farms[farm_name] = fname

            wb_obj.close()

    print('{} Farms checked OK for study {}'.format(len(farms), study) )

    return

def setup_sheet_data_dict(data_dict, var_format_dict):
    '''
    all classes require sheet_data dictionary to be initiated
    '''
    sheet_data = {}
    var_name_list = list(var_format_dict.keys())
    for var_name in var_name_list:
        sheet_data[var_name] = []

    sheet_data['period'] = data_dict['period']  # steady state or forward

    ntsteps = len(data_dict['period'])
    nyears = int(ntsteps/12)

    sheet_data['month'] = nyears * [tstep + 1 for tstep in range(12)]

    sheet_data['year'] = []
    this_year = -10
    for tstep in range(ntsteps):
        if int(tstep / 12) * 12 == tstep:
            if tstep > 0:
                this_year += 1

        sheet_data['year'].append(this_year)

    exclusion_list = list(['period', 'year', 'month'])

    return sheet_data, var_name_list, exclusion_list
